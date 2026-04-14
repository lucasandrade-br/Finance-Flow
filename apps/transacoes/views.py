from datetime import datetime
from decimal import Decimal, InvalidOperation
import csv
import difflib
import io
import re
import unicodedata

from django.db import transaction
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.clickjacking import xframe_options_sameorigin
from django.views.decorators.http import require_http_methods
import json

from apps.contas.models import ContaBancaria, PlanoConta, Tag
from apps.investimentos.models import AportePatrimonial
from apps.orcamento.models import Ciclo
from apps.transacoes.models import AliasImportacao, LancamentoFuturo, Movimentacao, MovimentacaoExcluida, TipoTransacao, TransacaoRecorrente


IMPORT_PREVIEW_SESSION_KEY = 'transacoes_import_preview'


def _add_months(data_base, meses):
	ano = data_base.year + ((data_base.month - 1 + meses) // 12)
	mes = ((data_base.month - 1 + meses) % 12) + 1
	# 28 evita problemas de virada de mês (29/30/31)
	dia = min(data_base.day, 28)
	return data_base.replace(year=ano, month=mes, day=dia)


def _to_decimal_or_zero(raw):
	try:
		return Decimal(str(raw or '0'))
	except (InvalidOperation, TypeError, ValueError):
		return Decimal('0')


def _parse_date_range(raw):
	if not raw:
		return (None, None)
	partes = re.findall(r'\d{4}-\d{2}-\d{2}', raw)
	if not partes:
		return (None, None)
	inicio = parse_date(partes[0])
	fim = parse_date(partes[-1]) if len(partes) > 1 else inicio
	if inicio and fim and inicio > fim:
		inicio, fim = fim, inicio
	return (inicio, fim)


def _normalizar_alias_texto(texto):
	valor = unicodedata.normalize('NFKD', (texto or '').strip())
	valor = ''.join(c for c in valor if not unicodedata.combining(c))
	return ' '.join(valor.lower().split())


def _resumo_destino_alias(alias):
	if alias.entidade == AliasImportacao.Entidade.TAG and alias.tag_id:
		return {'id': alias.tag_id, 'nome': alias.tag.nome}
	if alias.entidade == AliasImportacao.Entidade.PLANO_CONTA and alias.plano_conta_id:
		nome = f'{alias.plano_conta.codigo} - {alias.plano_conta.nome}' if alias.plano_conta.codigo else alias.plano_conta.nome
		return {'id': alias.plano_conta_id, 'nome': nome}
	if alias.entidade == AliasImportacao.Entidade.CONTA_BANCARIA and alias.conta_bancaria_id:
		return {'id': alias.conta_bancaria_id, 'nome': alias.conta_bancaria.nome}
	return {'id': None, 'nome': 'Sem vínculo'}


def _sugestoes_entidade(entidade, valor_externo):
	valor_normalizado = _normalizar_alias_texto(valor_externo)
	if not valor_normalizado:
		return []

	if entidade == AliasImportacao.Entidade.TAG:
		candidatos = [(tag.id, tag.nome, _normalizar_alias_texto(tag.nome)) for tag in Tag.objects.order_by('nome')]
	elif entidade == AliasImportacao.Entidade.PLANO_CONTA:
		candidatos = []
		for plano in PlanoConta.objects.order_by('codigo', 'nome'):
			nome_completo = f'{plano.codigo} - {plano.nome}' if plano.codigo else plano.nome
			candidatos.append((plano.id, nome_completo, _normalizar_alias_texto(nome_completo)))
	else:
		candidatos = [(conta.id, conta.nome, _normalizar_alias_texto(conta.nome)) for conta in ContaBancaria.objects.order_by('nome')]

	resultados = []
	for item_id, nome, nome_normalizado in candidatos:
		ratio = difflib.SequenceMatcher(None, valor_normalizado, nome_normalizado).ratio()
		if valor_normalizado in nome_normalizado or nome_normalizado in valor_normalizado:
			ratio = max(ratio, 0.9)
		if ratio >= 0.55:
			resultados.append({'id': item_id, 'nome': nome, 'score': round(ratio, 3)})

	resultados.sort(key=lambda item: item['score'], reverse=True)
	return resultados[:5]


def _normalizar_chave_coluna(valor):
	texto = _normalizar_alias_texto(valor)
	return re.sub(r'[^a-z0-9]+', '', texto)


def _chave_mapeamento_manual(valor):
	return _normalizar_chave_coluna(valor)


def _detectar_delimitador_csv(texto):
	try:
		amostra = texto[:4096]
		dialeto = csv.Sniffer().sniff(amostra, delimiters=',;\t|')
		return dialeto.delimiter
	except csv.Error:
		if texto.count(';') > texto.count(','):
			return ';'
		return ','


def _ler_csv_upload(arquivo):
	conteudo_bruto = arquivo.read()
	for encoding in ('utf-8-sig', 'latin-1'):
		try:
			conteudo = conteudo_bruto.decode(encoding)
			break
		except UnicodeDecodeError:
			conteudo = None
	if conteudo is None:
		raise ValueError('Não foi possível ler o arquivo CSV com UTF-8 ou Latin-1.')

	delimitador = _detectar_delimitador_csv(conteudo)
	leitor = csv.DictReader(io.StringIO(conteudo), delimiter=delimitador)
	headers = [h.strip() for h in (leitor.fieldnames or []) if (h or '').strip()]
	linhas = []
	for linha in leitor:
		linhas.append({(k or '').strip(): (v or '').strip() for k, v in linha.items()})
	return headers, linhas


def _ler_xlsx_upload(arquivo):
	try:
		from openpyxl import load_workbook
	except ImportError as exc:
		raise ValueError('Suporte a XLSX indisponível. Instale o pacote openpyxl.') from exc

	workbook = load_workbook(filename=io.BytesIO(arquivo.read()), data_only=True)
	sheet = workbook.active
	linhas_planilha = list(sheet.iter_rows(values_only=True))
	if not linhas_planilha:
		return [], []

	headers = []
	for celula in linhas_planilha[0]:
		headers.append(str(celula).strip() if celula is not None else '')
	headers = [h for h in headers if h]

	linhas = []
	for row_values in linhas_planilha[1:]:
		if row_values is None:
			continue
		row_dict = {}
		possui_valor = False
		for idx, header in enumerate(headers):
			valor = row_values[idx] if idx < len(row_values) else ''
			if valor is None:
				texto = ''
			elif hasattr(valor, 'strftime'):
				texto = valor.strftime('%Y-%m-%d')
			else:
				texto = str(valor).strip()
			if texto:
				possui_valor = True
			row_dict[header] = texto
		if possui_valor:
			linhas.append(row_dict)

	return headers, linhas


def _ler_arquivo_importacao(arquivo):
	nome = (arquivo.name or '').lower()
	if nome.endswith('.csv'):
		return _ler_csv_upload(arquivo)
	if nome.endswith('.xlsx'):
		return _ler_xlsx_upload(arquivo)
	raise ValueError('Formato inválido. Envie um arquivo CSV ou XLSX.')


def _inferir_mapeamento_colunas(headers):
	campos = {
		'descricao': ['descricao', 'descricao', 'historico', 'titulo', 'nome'],
		'tipo': ['tipo', 'natureza', 'entrada_saida', 'entrada/saida'],
		'valor': ['valor', 'valor_total', 'quantia', 'importe', 'amount'],
		'data_vencimento': ['data_vencimento', 'vencimento', 'data', 'competencia', 'dt_vencimento'],
		'data_pagamento': ['data_pagamento', 'pagamento', 'dt_pagamento'],
		'plano_conta': ['plano_conta', 'categoria', 'categoria_nome', 'plano', 'conta_contabil'],
		'conta_bancaria': ['conta_bancaria', 'conta', 'banco', 'conta_origem'],
		'tags': ['tags', 'tag', 'etiquetas', 'labels'],
		'status': ['status', 'situacao'],
		'formato_pagamento': ['formato_pagamento', 'forma_pagamento', 'forma', 'pagamento_forma'],
		'frequencia': ['frequencia', 'recorrencia'],
	}

	headers_norm = {header: _normalizar_chave_coluna(header) for header in headers}
	mapeamento = {}
	for campo, sinonimos in campos.items():
		sinonimos_norm = {_normalizar_chave_coluna(s) for s in sinonimos}
		escolhido = ''
		for header, header_norm in headers_norm.items():
			if header_norm in sinonimos_norm:
				escolhido = header
				break
		mapeamento[campo] = escolhido
	return mapeamento


def _parse_decimal_importacao(raw):
	texto = (raw or '').strip()
	if not texto:
		return Decimal('0')
	texto = texto.replace('R$', '').replace(' ', '')
	if ',' in texto and '.' in texto:
		if texto.rfind(',') > texto.rfind('.'):
			texto = texto.replace('.', '').replace(',', '.')
		else:
			texto = texto.replace(',', '')
	elif ',' in texto:
		texto = texto.replace('.', '').replace(',', '.')
	try:
		return Decimal(texto)
	except (InvalidOperation, TypeError, ValueError):
		return None


def _parse_data_importacao(raw):
	texto = (raw or '').strip()
	if not texto:
		return None
	for formato in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
		try:
			return datetime.strptime(texto, formato).date()
		except ValueError:
			continue
	tentativa = parse_date(texto)
	return tentativa


def _normalizar_tipo_importacao(raw, default_val):
	texto = _normalizar_alias_texto(raw)
	if not texto:
		return default_val
	mapeamento = {
		'receita': TipoTransacao.RECEITA,
		'despesa': TipoTransacao.DESPESA,
		'investimento': TipoTransacao.INVESTIMENTO,
		'transferencia': TipoTransacao.TRANSFERENCIA,
		'transfentrada': TipoTransacao.TRANSFERENCIA_ENTRADA,
		'transfsaida': TipoTransacao.TRANSFERENCIA_SAIDA,
	}
	compactado = texto.replace(' ', '')
	if compactado in mapeamento:
		return mapeamento[compactado]
	for chave, valor in mapeamento.items():
		if chave in compactado:
			return valor
	return default_val


def _normalizar_choice_importacao(raw, choices, default_val):
	texto = _normalizar_alias_texto(raw)
	if not texto:
		return default_val
	for value, label in choices:
		if texto in (_normalizar_alias_texto(value), _normalizar_alias_texto(label)):
			return value
	return default_val


def _resolver_por_entidade(entidade, valor_raw, manual_map):
	valor_limpo = ' '.join((valor_raw or '').strip().split())
	if not valor_limpo:
		return (None, 'vazio')

	chave_manual = _chave_mapeamento_manual(valor_limpo)
	if chave_manual in manual_map:
		return (manual_map[chave_manual], 'manual')

	alias = AliasImportacao.objects.filter(
		entidade=entidade,
		valor_externo__iexact=valor_limpo,
		ativo=True,
	).first()
	if alias:
		if entidade == AliasImportacao.Entidade.TAG and alias.tag_id:
			return (alias.tag_id, 'alias')
		if entidade == AliasImportacao.Entidade.PLANO_CONTA and alias.plano_conta_id:
			return (alias.plano_conta_id, 'alias')
		if entidade == AliasImportacao.Entidade.CONTA_BANCARIA and alias.conta_bancaria_id:
			return (alias.conta_bancaria_id, 'alias')

	valor_norm = _normalizar_alias_texto(valor_limpo)
	if entidade == AliasImportacao.Entidade.TAG:
		for tag in Tag.objects.order_by('nome'):
			if _normalizar_alias_texto(tag.nome) == valor_norm:
				return (tag.id, 'exato')
	elif entidade == AliasImportacao.Entidade.PLANO_CONTA:
		for plano in PlanoConta.objects.order_by('codigo', 'nome'):
			nome_completo = f'{plano.codigo} - {plano.nome}' if plano.codigo else plano.nome
			if _normalizar_alias_texto(nome_completo) == valor_norm or _normalizar_alias_texto(plano.nome) == valor_norm:
				return (plano.id, 'exato')
	else:
		for conta in ContaBancaria.objects.order_by('nome'):
			if _normalizar_alias_texto(conta.nome) == valor_norm:
				return (conta.id, 'exato')

	sugestoes = _sugestoes_entidade(entidade, valor_limpo)
	if sugestoes and sugestoes[0]['score'] >= 0.9:
		return (sugestoes[0]['id'], 'sugestao')

	return (None, 'nao_encontrado')


def _montar_preview_importacao(payload, manual_maps=None):
	manual_maps = manual_maps or {
		'plano': {},
		'conta': {},
		'tag': {},
	}

	mapeamento = payload.get('mapeamento', {})
	defaults = payload.get('defaults', {})
	linhas_brutas = payload.get('linhas', [])

	preview_rows = []
	unresolved = {
		'plano': {},
		'conta': {},
		'tag': {},
	}

	for indice, row in enumerate(linhas_brutas, start=1):
		erros = []
		descricao = (row.get(mapeamento.get('descricao') or '', '') or '').strip() or f'Importação linha {indice}'
		valor_raw = row.get(mapeamento.get('valor') or '', '')
		valor = _parse_decimal_importacao(valor_raw)
		if valor is None:
			erros.append('Valor inválido')

		data_vencimento_raw = row.get(mapeamento.get('data_vencimento') or '', '')
		data_pagamento_raw = row.get(mapeamento.get('data_pagamento') or '', '')
		data_vencimento = _parse_data_importacao(data_vencimento_raw)
		data_pagamento = _parse_data_importacao(data_pagamento_raw)
		if not data_vencimento:
			if data_pagamento:
				data_vencimento = data_pagamento
			else:
				erros.append('Data de vencimento inválida')

		tipo = _normalizar_tipo_importacao(
			row.get(mapeamento.get('tipo') or '', ''),
			defaults.get('tipo') or TipoTransacao.DESPESA,
		)
		status = _normalizar_choice_importacao(
			row.get(mapeamento.get('status') or '', ''),
			Movimentacao.Status.choices,
			defaults.get('status') or Movimentacao.Status.PENDENTE,
		)
		formato_pagamento = _normalizar_choice_importacao(
			row.get(mapeamento.get('formato_pagamento') or '', ''),
			Movimentacao._meta.get_field('formato_pagamento').choices,
			defaults.get('formato_pagamento') or 'PIX',
		)
		frequencia = _normalizar_choice_importacao(
			row.get(mapeamento.get('frequencia') or '', ''),
			Movimentacao._meta.get_field('frequencia').choices,
			defaults.get('frequencia') or 'Variavel',
		)

		plano_raw = row.get(mapeamento.get('plano_conta') or '', '')
		conta_raw = row.get(mapeamento.get('conta_bancaria') or '', '')
		tags_raw = row.get(mapeamento.get('tags') or '', '')

		plano_id, plano_origem = _resolver_por_entidade(AliasImportacao.Entidade.PLANO_CONTA, plano_raw, manual_maps.get('plano', {}))
		conta_id, conta_origem = _resolver_por_entidade(AliasImportacao.Entidade.CONTA_BANCARIA, conta_raw, manual_maps.get('conta', {}))

		tag_ids = []
		tags_pendentes = []
		for raw_tag in [t.strip() for t in re.split(r'[;,|]', tags_raw or '') if t.strip()]:
			tag_id, tag_origem = _resolver_por_entidade(AliasImportacao.Entidade.TAG, raw_tag, manual_maps.get('tag', {}))
			if tag_id:
				tag_ids.append(tag_id)
			else:
				tags_pendentes.append(raw_tag)
				chave_tag = _chave_mapeamento_manual(raw_tag)
				if chave_tag not in unresolved['tag']:
					unresolved['tag'][chave_tag] = {
						'raw': raw_tag,
						'sugestoes': _sugestoes_entidade(AliasImportacao.Entidade.TAG, raw_tag),
					}

		if plano_raw and not plano_id:
			chave_plano = _chave_mapeamento_manual(plano_raw)
			if chave_plano not in unresolved['plano']:
				unresolved['plano'][chave_plano] = {
					'raw': plano_raw,
					'sugestoes': _sugestoes_entidade(AliasImportacao.Entidade.PLANO_CONTA, plano_raw),
				}
			erros.append('Plano de conta não resolvido')

		if conta_raw and not conta_id:
			chave_conta = _chave_mapeamento_manual(conta_raw)
			if chave_conta not in unresolved['conta']:
				unresolved['conta'][chave_conta] = {
					'raw': conta_raw,
					'sugestoes': _sugestoes_entidade(AliasImportacao.Entidade.CONTA_BANCARIA, conta_raw),
				}
			erros.append('Conta bancária não resolvida')

		if tags_pendentes:
			erros.append('Tags não resolvidas')

		preview_rows.append({
			'linha': indice,
			'descricao': descricao,
			'tipo': tipo,
			'valor': valor,
			'data_vencimento': data_vencimento,
			'data_pagamento': data_pagamento,
			'plano_conta_id': plano_id,
			'conta_bancaria_id': conta_id,
			'tag_ids': sorted(set(tag_ids)),
			'status': status,
			'formato_pagamento': formato_pagamento,
			'frequencia': frequencia,
			'origens': {
				'plano': plano_origem,
				'conta': conta_origem,
			},
			'raw': {
				'plano': plano_raw,
				'conta': conta_raw,
				'tags': tags_raw,
			},
			'erros': erros,
		})

	total_erros = sum(1 for item in preview_rows if item['erros'])
	return {
		'rows': preview_rows,
		'unresolved': unresolved,
		'total_rows': len(preview_rows),
		'total_invalidas': total_erros,
	}


def _coletar_mapeamento_manual(post_data):
	resultado = {'plano': {}, 'conta': {}, 'tag': {}}
	prefixos = {
		'manual_plano_': 'plano',
		'manual_conta_': 'conta',
		'manual_tag_': 'tag',
	}
	for chave, valor in post_data.items():
		for prefixo, destino in prefixos.items():
			if chave.startswith(prefixo) and valor:
				key = chave.replace(prefixo, '', 1)
				try:
					resultado[destino][key] = int(valor)
				except (TypeError, ValueError):
					pass
	return resultado


def _persistir_aliases_manuais(preview_rows, manual_maps):
	for row in preview_rows:
		plano_raw = row['raw'].get('plano')
		conta_raw = row['raw'].get('conta')
		tags_raw = row['raw'].get('tags')

		if plano_raw:
			chave = _chave_mapeamento_manual(plano_raw)
			plano_id = manual_maps.get('plano', {}).get(chave)
			if plano_id:
				AliasImportacao.objects.get_or_create(
					entidade=AliasImportacao.Entidade.PLANO_CONTA,
					valor_externo=plano_raw,
					defaults={'plano_conta_id': plano_id, 'ativo': True},
				)

		if conta_raw:
			chave = _chave_mapeamento_manual(conta_raw)
			conta_id = manual_maps.get('conta', {}).get(chave)
			if conta_id:
				AliasImportacao.objects.get_or_create(
					entidade=AliasImportacao.Entidade.CONTA_BANCARIA,
					valor_externo=conta_raw,
					defaults={'conta_bancaria_id': conta_id, 'ativo': True},
				)

		for raw_tag in [t.strip() for t in re.split(r'[;,|]', tags_raw or '') if t.strip()]:
			chave = _chave_mapeamento_manual(raw_tag)
			tag_id = manual_maps.get('tag', {}).get(chave)
			if tag_id:
				AliasImportacao.objects.get_or_create(
					entidade=AliasImportacao.Entidade.TAG,
					valor_externo=raw_tag,
					defaults={'tag_id': tag_id, 'ativo': True},
				)


def livro_razao(request):
	data_range = (request.GET.get('data_range') or '').strip()
	descricao = (request.GET.get('descricao') or '').strip()
	valor_operador = (request.GET.get('valor_operador') or '').strip()
	valor_filtro = (request.GET.get('valor_filtro') or '').strip()
	tipo = (request.GET.get('tipo') or '').strip()
	plano_conta_id = (request.GET.get('plano_conta_id') or '').strip()
	conta_bancaria_id = (request.GET.get('conta_bancaria_id') or '').strip()
	if 'status' in request.GET:
		status = (request.GET.get('status') or '').strip()
	else:
		status = Movimentacao.Status.VALIDADO
	tag_id = (request.GET.get('tag_id') or '').strip()
	per_page_raw = (request.GET.get('per_page') or '30').strip()
	try:
		per_page = int(per_page_raw)
	except (TypeError, ValueError):
		per_page = 30
	if per_page not in (30, 50, 100):
		per_page = 30

	transacoes = Movimentacao.objects.select_related(
		'plano_conta',
		'conta_bancaria',
	).prefetch_related('tags').annotate(
		data_referencia=Coalesce('data_pagamento', 'data_vencimento'),
	)

	if data_range:
		data_inicio, data_fim = _parse_date_range(data_range)
		if data_inicio and data_fim:
			transacoes = transacoes.filter(data_referencia__range=(data_inicio, data_fim))
	if descricao:
		transacoes = transacoes.filter(descricao__icontains=descricao)
	if valor_filtro and valor_operador in ('gt', 'lt', 'eq'):
		valor_normalizado = valor_filtro.replace('.', '').replace(',', '.') if ',' in valor_filtro else valor_filtro
		try:
			valor_decimal = Decimal(valor_normalizado)
			if valor_operador == 'gt':
				transacoes = transacoes.filter(valor__gt=valor_decimal)
			elif valor_operador == 'lt':
				transacoes = transacoes.filter(valor__lt=valor_decimal)
			else:
				transacoes = transacoes.filter(valor=valor_decimal)
		except (InvalidOperation, TypeError, ValueError):
			pass
	if tipo:
		transacoes = transacoes.filter(tipo=tipo)
	if plano_conta_id:
		transacoes = transacoes.filter(plano_conta_id=plano_conta_id)
	if conta_bancaria_id:
		transacoes = transacoes.filter(conta_bancaria_id=conta_bancaria_id)
	if status:
		transacoes = transacoes.filter(status=status)
	if tag_id:
		transacoes = transacoes.filter(tags__id=tag_id)

	transacoes = transacoes.distinct().order_by('-data_referencia', '-created_at')
	paginator = Paginator(transacoes, per_page)
	page_number = request.GET.get('page')
	page_obj = paginator.get_page(page_number)

	query_params = request.GET.copy()
	if 'page' in query_params:
		query_params.pop('page')
	query_string = query_params.urlencode()
	page_query_prefix = f'{query_string}&' if query_string else ''

	movimentacoes_excluidas = MovimentacaoExcluida.objects.select_related(
		'plano_conta',
		'conta_bancaria',
	).order_by('-excluida_em', '-created_at')[:100]

	context = {
		'transacoes': page_obj,
		'movimentacoes_excluidas': movimentacoes_excluidas,
		'tipos_transacao': TipoTransacao.choices,
		'status_choices': Movimentacao.Status.choices,
		'plano_contas': PlanoConta.objects.order_by('codigo', 'nome'),
		'contas_bancarias': ContaBancaria.objects.order_by('nome'),
		'tags': Tag.objects.order_by('nome'),
		'page_query_prefix': page_query_prefix,
		'filtros': {
			'data_range': data_range,
			'descricao': descricao,
			'valor_operador': valor_operador,
			'valor_filtro': valor_filtro,
			'tipo': tipo,
			'plano_conta_id': plano_conta_id,
			'conta_bancaria_id': conta_bancaria_id,
			'status': status,
			'tag_id': tag_id,
			'per_page': str(per_page),
		},
	}
	return render(request, 'transacoes/padrao/livro_razao.html', context)


def lista_aliases_importacao(request):
	entidade = (request.GET.get('entidade') or '').strip()
	busca = (request.GET.get('q') or '').strip()

	aliases = AliasImportacao.objects.select_related('tag', 'plano_conta', 'conta_bancaria').order_by('entidade', 'valor_externo')
	if entidade:
		aliases = aliases.filter(entidade=entidade)
	if busca:
		aliases = aliases.filter(valor_externo__icontains=busca)

	for alias in aliases:
		alias.destino_resumo = _resumo_destino_alias(alias)

	context = {
		'aliases': aliases,
		'entidades': AliasImportacao.Entidade.choices,
		'filtros': {
			'entidade': entidade,
			'q': busca,
		},
	}
	return render(request, 'transacoes/importacao/lista_aliases.html', context)


def novo_alias_importacao(request):
	entidade = request.POST.get('entidade') or request.GET.get('entidade') or AliasImportacao.Entidade.TAG
	erro_formulario = None
	alias = None
	alias_id = request.GET.get('alias_id')
	if alias_id:
		alias = get_object_or_404(AliasImportacao, pk=alias_id)
		entidade = alias.entidade

	if request.method == 'POST':
		alias_id_post = request.POST.get('alias_id')
		if alias_id_post:
			alias = get_object_or_404(AliasImportacao, pk=alias_id_post)

		entidade = request.POST.get('entidade') or entidade
		valor_externo = (request.POST.get('valor_externo') or '').strip()
		valor_externo = ' '.join(valor_externo.split())
		tag_id = request.POST.get('tag_id') or None
		plano_conta_id = request.POST.get('plano_conta_id') or None
		conta_bancaria_id = request.POST.get('conta_bancaria_id') or None
		ativo = bool(request.POST.get('ativo'))

		if not valor_externo:
			erro_formulario = 'Informe o valor externo como aparece na planilha.'
		else:
			existing_qs = AliasImportacao.objects.filter(entidade=entidade, valor_externo=valor_externo)
			if alias:
				existing_qs = existing_qs.exclude(pk=alias.pk)

			if existing_qs.exists():
				erro_formulario = 'Já existe um alias para esse valor externo nesta entidade.'
			elif entidade == AliasImportacao.Entidade.TAG and not tag_id:
				erro_formulario = 'Selecione a Tag de destino.'
			elif entidade == AliasImportacao.Entidade.PLANO_CONTA and not plano_conta_id:
				erro_formulario = 'Selecione o Plano de Conta de destino.'
			elif entidade == AliasImportacao.Entidade.CONTA_BANCARIA and not conta_bancaria_id:
				erro_formulario = 'Selecione a Conta Bancária de destino.'

		if not erro_formulario:
			if not alias:
				alias = AliasImportacao()
			alias.entidade = entidade
			alias.valor_externo = valor_externo
			alias.tag_id = tag_id if entidade == AliasImportacao.Entidade.TAG else None
			alias.plano_conta_id = plano_conta_id if entidade == AliasImportacao.Entidade.PLANO_CONTA else None
			alias.conta_bancaria_id = conta_bancaria_id if entidade == AliasImportacao.Entidade.CONTA_BANCARIA else None
			alias.ativo = ativo
			alias.save()
			return redirect('transacoes:lista_aliases_importacao')

	context = {
		'alias': alias,
		'entidade': entidade,
		'entidades': AliasImportacao.Entidade.choices,
		'tags': Tag.objects.order_by('nome'),
		'planos_conta': PlanoConta.objects.order_by('codigo', 'nome'),
		'contas_bancarias': ContaBancaria.objects.order_by('nome'),
		'erro_formulario': erro_formulario,
	}
	return render(request, 'transacoes/importacao/form_alias.html', context)


@require_http_methods(["POST"])
def excluir_alias_importacao(request, alias_id):
	alias = get_object_or_404(AliasImportacao, pk=alias_id)
	alias.delete()
	return redirect('transacoes:lista_aliases_importacao')


def sugerir_correspondencia_importacao(request):
	entidade = (request.GET.get('entidade') or '').strip()
	valor_externo = (request.GET.get('valor') or '').strip()

	if entidade not in dict(AliasImportacao.Entidade.choices):
		return JsonResponse({'error': 'Entidade inválida.'}, status=400)

	alias = AliasImportacao.objects.select_related('tag', 'plano_conta', 'conta_bancaria').filter(
		entidade=entidade,
		valor_externo__iexact=valor_externo,
		ativo=True,
	).first()

	if alias:
		return JsonResponse({
			'match_type': 'alias',
			'alias_id': alias.id,
			'destino': _resumo_destino_alias(alias),
			'sugestoes': [],
		})

	sugestoes = _sugestoes_entidade(entidade, valor_externo)
	return JsonResponse({
		'match_type': 'sugestao' if sugestoes else 'nao_encontrado',
		'destino': None,
		'sugestoes': sugestoes,
	})


def importar_transacoes(request):
	payload = request.session.get(IMPORT_PREVIEW_SESSION_KEY)
	preview = _montar_preview_importacao(payload) if payload else None
	erro_formulario = None

	headers = payload.get('headers', []) if payload else []
	mapeamento = payload.get('mapeamento', _inferir_mapeamento_colunas(headers)) if payload else {}
	defaults = payload.get('defaults', {
		'tipo': TipoTransacao.DESPESA,
		'status': Movimentacao.Status.PENDENTE,
		'formato_pagamento': 'PIX',
		'frequencia': 'Variavel',
	}) if payload else {
		'tipo': TipoTransacao.DESPESA,
		'status': Movimentacao.Status.PENDENTE,
		'formato_pagamento': 'PIX',
		'frequencia': 'Variavel',
	}

	if request.method == 'POST':
		action = request.POST.get('action') or 'preview'
		if action == 'preview':
			arquivo = request.FILES.get('arquivo')
			if not arquivo:
				erro_formulario = 'Selecione um arquivo CSV ou XLSX para gerar o preview.'
			else:
				try:
					headers, linhas = _ler_arquivo_importacao(arquivo)
				except ValueError as exc:
					erro_formulario = str(exc)
					linhas = []
					headers = []

				if not erro_formulario:
					if not headers:
						erro_formulario = 'Arquivo sem cabeçalho. Inclua a linha de títulos das colunas.'
					elif not linhas:
						erro_formulario = 'Arquivo sem linhas de dados para importar.'
					elif len(linhas) > 2000:
						erro_formulario = 'Limite de 2000 linhas por importação para manter a conferência segura.'

				if not erro_formulario:
					mapeamento_auto = _inferir_mapeamento_colunas(headers)
					mapeamento = {
						'descricao': request.POST.get('col_descricao') or mapeamento_auto.get('descricao') or '',
						'tipo': request.POST.get('col_tipo') or mapeamento_auto.get('tipo') or '',
						'valor': request.POST.get('col_valor') or mapeamento_auto.get('valor') or '',
						'data_vencimento': request.POST.get('col_data_vencimento') or mapeamento_auto.get('data_vencimento') or '',
						'data_pagamento': request.POST.get('col_data_pagamento') or mapeamento_auto.get('data_pagamento') or '',
						'plano_conta': request.POST.get('col_plano_conta') or mapeamento_auto.get('plano_conta') or '',
						'conta_bancaria': request.POST.get('col_conta_bancaria') or mapeamento_auto.get('conta_bancaria') or '',
						'tags': request.POST.get('col_tags') or mapeamento_auto.get('tags') or '',
						'status': request.POST.get('col_status') or mapeamento_auto.get('status') or '',
						'formato_pagamento': request.POST.get('col_formato_pagamento') or mapeamento_auto.get('formato_pagamento') or '',
						'frequencia': request.POST.get('col_frequencia') or mapeamento_auto.get('frequencia') or '',
					}

					if not mapeamento.get('valor'):
						erro_formulario = 'Mapeie a coluna de valor para continuar.'
					elif not mapeamento.get('data_vencimento') and not mapeamento.get('data_pagamento'):
						erro_formulario = 'Mapeie ao menos data de vencimento ou data de pagamento.'

				if not erro_formulario:
					defaults = {
						'tipo': request.POST.get('default_tipo') or TipoTransacao.DESPESA,
						'status': request.POST.get('default_status') or Movimentacao.Status.PENDENTE,
						'formato_pagamento': request.POST.get('default_formato_pagamento') or 'PIX',
						'frequencia': request.POST.get('default_frequencia') or 'Variavel',
					}
					payload = {
						'headers': headers,
						'linhas': linhas,
						'mapeamento': mapeamento,
						'defaults': defaults,
					}
					request.session[IMPORT_PREVIEW_SESSION_KEY] = payload
					request.session.modified = True
					preview = _montar_preview_importacao(payload)

	context = {
		'erro_formulario': erro_formulario,
		'headers': headers,
		'mapeamento': mapeamento,
		'defaults': defaults,
		'preview': preview,
		'tipos_transacao': TipoTransacao.choices,
		'status_choices': Movimentacao.Status.choices,
		'formato_choices': Movimentacao._meta.get_field('formato_pagamento').choices,
		'frequencia_choices': Movimentacao._meta.get_field('frequencia').choices,
		'planos_conta': PlanoConta.objects.order_by('codigo', 'nome'),
		'contas_bancarias': ContaBancaria.objects.order_by('nome'),
		'tags': Tag.objects.order_by('nome'),
	}
	return render(request, 'transacoes/importacao/importar_transacoes.html', context)


@require_http_methods(["POST"])
def confirmar_importacao_transacoes(request):
	payload = request.session.get(IMPORT_PREVIEW_SESSION_KEY)
	if not payload:
		return redirect('transacoes:importar_transacoes')

	manual_maps = _coletar_mapeamento_manual(request.POST)
	criar_tags_nao_localizadas = bool(request.POST.get('criar_tags_nao_localizadas'))

	if criar_tags_nao_localizadas:
		preview_base = _montar_preview_importacao(payload, manual_maps=manual_maps)
		for key, item in preview_base.get('unresolved', {}).get('tag', {}).items():
			nome_tag = (item.get('raw') or '').strip()[:100]
			if not nome_tag:
				continue
			tag = Tag.objects.filter(nome__iexact=nome_tag).first()
			if not tag:
				tag = Tag.objects.create(nome=nome_tag)
			manual_maps.setdefault('tag', {})[key] = tag.id

	preview = _montar_preview_importacao(payload, manual_maps=manual_maps)
	if preview['total_invalidas'] > 0:
		context = {
			'erro_formulario': 'Ainda existem linhas pendentes de reconciliação. Ajuste os mapeamentos manuais e tente confirmar novamente.',
			'headers': payload.get('headers', []),
			'mapeamento': payload.get('mapeamento', {}),
			'defaults': payload.get('defaults', {}),
			'preview': preview,
			'tipos_transacao': TipoTransacao.choices,
			'status_choices': Movimentacao.Status.choices,
			'formato_choices': Movimentacao._meta.get_field('formato_pagamento').choices,
			'frequencia_choices': Movimentacao._meta.get_field('frequencia').choices,
			'planos_conta': PlanoConta.objects.order_by('codigo', 'nome'),
			'contas_bancarias': ContaBancaria.objects.order_by('nome'),
			'tags': Tag.objects.order_by('nome'),
		}
		return render(request, 'transacoes/importacao/importar_transacoes.html', context)

	criar_aliases = bool(request.POST.get('criar_aliases'))
	with transaction.atomic():
		for row in preview['rows']:
			movimentacao = Movimentacao.objects.create(
				tipo=row['tipo'],
				valor=row['valor'],
				formato_pagamento=row['formato_pagamento'],
				frequencia=row['frequencia'],
				descricao=row['descricao'],
				data_pagamento=row['data_pagamento'],
				data_vencimento=row['data_vencimento'] or timezone.localdate(),
				plano_conta_id=row['plano_conta_id'],
				conta_bancaria_id=row['conta_bancaria_id'],
				status=row['status'],
			)
			if row['tag_ids']:
				movimentacao.tags.set(row['tag_ids'])

		if criar_aliases:
			_persistir_aliases_manuais(preview['rows'], manual_maps)

	request.session.pop(IMPORT_PREVIEW_SESSION_KEY, None)
	return redirect('transacoes:livro_razao')


def baixar_modelo_importacao_xlsx(request):
	try:
		from openpyxl import Workbook
	except ImportError:
		return JsonResponse(
			{'error': 'Biblioteca openpyxl não encontrada para gerar o modelo XLSX.'},
			status=500,
		)

	workbook = Workbook()
	sheet = workbook.active
	sheet.title = 'importacao'

	headers = [
		'descricao',
		'tipo',
		'valor',
		'data_vencimento',
		'data_pagamento',
		'plano_conta',
		'conta_bancaria',
		'tags',
		'status',
		'formato_pagamento',
		'frequencia',
	]
	sheet.append(headers)
	sheet.append([
		'Restaurante Tio Nado',
		'Despesa',
		'89,90',
		'2026-04-05',
		'2026-04-05',
		'Alimentação',
		'Inter PJ',
		'almoço;cliente',
		'Pendente',
		'PIX',
		'Variavel',
	])
	sheet.append([
		'Recebimento Projeto X',
		'Receita',
		'2500.00',
		'2026-04-10',
		'2026-04-10',
		'Receitas de Serviços',
		'Nubank PJ',
		'cliente x;serviços',
		'Efetivado',
		'TED',
		'Variavel',
	])

	dominios = workbook.create_sheet('dominios')
	dominios.append(['campo', 'valores_aceitos'])
	dominios.append(['tipo', ', '.join([v for v, _ in TipoTransacao.choices])])
	dominios.append(['status', ', '.join([v for v, _ in Movimentacao.Status.choices])])
	dominios.append(['formato_pagamento', ', '.join([v for v, _ in Movimentacao._meta.get_field('formato_pagamento').choices])])
	dominios.append(['frequencia', ', '.join([v for v, _ in Movimentacao._meta.get_field('frequencia').choices])])
	dominios.append(['datas', 'Use preferencialmente YYYY-MM-DD'])
	dominios.append(['tags', 'Separar múltiplas tags por ; ou ,'])

	conteudo = io.BytesIO()
	workbook.save(conteudo)
	conteudo.seek(0)

	response = HttpResponse(
		conteudo.getvalue(),
		content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
	)
	response['Content-Disposition'] = 'attachment; filename="modelo_importacao_movimentacoes.xlsx"'
	return response


@require_http_methods(["POST"])
def excluir_movimentacao(request, movimentacao_id):
	movimentacao = get_object_or_404(Movimentacao, pk=movimentacao_id)
	next_url = request.POST.get('next') or request.GET.get('next')
	if next_url and not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
		next_url = None

	tags_snapshot = list(movimentacao.tags.values('id', 'nome'))

	with transaction.atomic():
		MovimentacaoExcluida.objects.create(
			original_movimentacao_id=movimentacao.id,
			tipo=movimentacao.tipo,
			valor=movimentacao.valor,
			formato_pagamento=movimentacao.formato_pagamento,
			frequencia=movimentacao.frequencia,
			descricao=movimentacao.descricao,
			comprovante_path=movimentacao.comprovante.name if movimentacao.comprovante else '',
			data_pagamento=movimentacao.data_pagamento,
			data_vencimento=movimentacao.data_vencimento,
			parcela_atual=movimentacao.parcela_atual,
			total_parcelas=movimentacao.total_parcelas,
			plano_conta=movimentacao.plano_conta,
			conta_bancaria=movimentacao.conta_bancaria,
			conta_destino=movimentacao.conta_destino,
			ciclo=movimentacao.ciclo,
			cofre=movimentacao.cofre,
			status=movimentacao.status,
			tags_snapshot=tags_snapshot,
		)
		movimentacao.delete()

	return redirect(next_url or 'transacoes:livro_razao')


@require_http_methods(["POST"])
def restaurar_movimentacao_excluida(request, item_id):
	item = get_object_or_404(MovimentacaoExcluida, pk=item_id)
	next_url = request.POST.get('next') or request.GET.get('next')
	if next_url and not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
		next_url = None

	with transaction.atomic():
		movimentacao = Movimentacao.objects.create(
			tipo=item.tipo,
			valor=item.valor,
			formato_pagamento=item.formato_pagamento,
			frequencia=item.frequencia,
			descricao=item.descricao,
			data_pagamento=item.data_pagamento,
			data_vencimento=item.data_vencimento or timezone.localdate(),
			parcela_atual=item.parcela_atual,
			total_parcelas=item.total_parcelas,
			plano_conta=item.plano_conta,
			conta_bancaria=item.conta_bancaria,
			conta_destino=item.conta_destino,
			ciclo=item.ciclo,
			cofre=item.cofre,
			status=item.status,
		)
		if item.comprovante_path:
			movimentacao.comprovante = item.comprovante_path
			movimentacao.save(update_fields=['comprovante', 'updated_at'])

		tag_ids = [registro.get('id') for registro in (item.tags_snapshot or []) if registro.get('id')]
		if tag_ids:
			movimentacao.tags.set(Tag.objects.filter(id__in=tag_ids))

		item.delete()

	return redirect(next_url or 'transacoes:livro_razao')


def nova_transacao(request):
	contas = ContaBancaria.objects.all().order_by('nome')
	plano_contas = PlanoConta.objects.all().order_by('codigo', 'nome')
	tags = Tag.objects.all().order_by('nome')
	ciclo_id_raw = request.POST.get('ciclo_id') or request.GET.get('ciclo_id')
	next_url = request.POST.get('next') or request.GET.get('next')

	if next_url and not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
		next_url = None

	ciclo_id = None
	if ciclo_id_raw:
		try:
			ciclo_id = int(ciclo_id_raw)
		except (TypeError, ValueError):
			ciclo_id = None

	def _build_form_context(form_data=None, erro_formulario=None):
		form_data = form_data or {}
		tag_ids = [str(tag_id) for tag_id in (form_data.get('tag_ids') or []) if str(tag_id)]
		selected_tags = list(
			Tag.objects.filter(id__in=tag_ids)
			.values('id', 'nome', 'plano_conta_id')
		)
		selected_tags_data = [
			{
				'id': str(item['id']),
				'text': item['nome'],
				'plano_conta_id': item['plano_conta_id'],
			}
			for item in selected_tags
		]

		return {
			'contas': contas,
			'plano_contas': plano_contas,
			'tags': tags,
			'hoje': timezone.localdate(),
			'ciclo_id_form': ciclo_id,
			'next_url_form': next_url,
			'cancel_url': next_url,
			'status_default_form': form_data.get('status') or (Movimentacao.Status.PENDENTE if ciclo_id else Movimentacao.Status.EFETIVADO),
			'tipo_default_form': form_data.get('tipo') or TipoTransacao.DESPESA,
			'form_data': form_data,
			'selected_tags_data': selected_tags_data,
			'erro_formulario': erro_formulario,
			'sincronizar_investimento_form': bool(form_data.get('sincronizar_investimento')),
		}

	if request.method == 'POST':
		tipo = request.POST.get('tipo')
		valor = request.POST.get('valor')
		descricao = request.POST.get('descricao', '').strip()
		data_pagamento = request.POST.get('data_pagamento')
		conta_bancaria_id = request.POST.get('conta_bancaria_id')
		conta_destino_id = request.POST.get('conta_destino_id')
		plano_conta_id = request.POST.get('plano_conta_id') or request.POST.get('plano_contas_id')
		status = request.POST.get('status')
		formato_pagamento = request.POST.get('formato_pagamento', 'PIX')
		frequencia = request.POST.get('frequencia', 'Variavel')
		comprovante = request.FILES.get('comprovante')
		tag_ids = request.POST.getlist('tags')
		sincronizar_investimento = request.POST.get('sincronizar_investimento') == 'on'

		form_data = {
			'tipo': tipo,
			'valor': valor,
			'descricao': descricao,
			'data_pagamento': data_pagamento,
			'conta_bancaria_id': conta_bancaria_id,
			'conta_destino_id': conta_destino_id,
			'plano_conta_id': plano_conta_id,
			'status': status,
			'formato_pagamento': formato_pagamento,
			'frequencia': frequencia,
			'tag_ids': tag_ids,
			'sincronizar_investimento': sincronizar_investimento,
		}

		required_fields = [tipo, valor, data_pagamento, conta_bancaria_id, plano_conta_id, status]
		if any(not field for field in required_fields):
			context = _build_form_context(form_data=form_data, erro_formulario='Ainda faltam dados obrigatorios. Revise os campos destacados e tente novamente.')
			return render(request, 'transacoes/padrao/form_lancamento.html', context)

		if tipo == TipoTransacao.TRANSFERENCIA:
			if not conta_destino_id or conta_destino_id == conta_bancaria_id:
				context = _build_form_context(
					form_data=form_data,
					erro_formulario='Selecione uma conta destino diferente da conta de origem para transferencias.',
				)
				return render(request, 'transacoes/padrao/form_lancamento.html', context)

			with transaction.atomic():
				mov_saida = Movimentacao.objects.create(
					tipo=TipoTransacao.TRANSFERENCIA_SAIDA,
					valor=valor,
					descricao=descricao,
					data_pagamento=data_pagamento,
					data_vencimento=data_pagamento or timezone.now().date(),
					conta_bancaria_id=conta_bancaria_id,
					conta_destino_id=conta_destino_id,
					plano_conta_id=plano_conta_id,
					status=status,
					formato_pagamento=formato_pagamento,
					frequencia=frequencia,
					comprovante=comprovante,
					ciclo_id=ciclo_id,
				)
				mov_entrada = Movimentacao.objects.create(
					tipo=TipoTransacao.TRANSFERENCIA_ENTRADA,
					valor=valor,
					descricao=descricao,
					data_pagamento=data_pagamento,
					data_vencimento=data_pagamento or timezone.now().date(),
					conta_bancaria_id=conta_destino_id,
					conta_destino_id=conta_bancaria_id,
					plano_conta_id=plano_conta_id,
					status=status,
					formato_pagamento=formato_pagamento,
					frequencia=frequencia,
					ciclo_id=ciclo_id,
				)
				mov_saida.lancamento_par = mov_entrada
				mov_saida.save(update_fields=['lancamento_par', 'updated_at'])
				mov_entrada.lancamento_par = mov_saida
				mov_entrada.save(update_fields=['lancamento_par', 'updated_at'])

				if tag_ids:
					mov_saida.tags.set(tag_ids)
					mov_entrada.tags.set(tag_ids)

			if next_url:
				return redirect(next_url)

			return redirect('transacoes:livro_razao')

		movimentacao = Movimentacao.objects.create(
			tipo=tipo,
			valor=valor,
			descricao=descricao,
			data_pagamento=data_pagamento,
			data_vencimento=data_pagamento or timezone.now().date(),
			conta_bancaria_id=conta_bancaria_id,
			conta_destino_id=conta_destino_id or None,
			plano_conta_id=plano_conta_id,
			status=status,
			formato_pagamento=formato_pagamento,
			frequencia=frequencia,
			comprovante=comprovante,
			ciclo_id=ciclo_id,
		)

		if tag_ids:
			movimentacao.tags.set(tag_ids)

		if tipo == TipoTransacao.INVESTIMENTO and sincronizar_investimento:
			plano_conta = PlanoConta.objects.filter(pk=plano_conta_id).only('codigo', 'nome').first()
			if plano_conta and plano_conta.codigo == '3.1.1' and 'renda' in (plano_conta.nome or '').lower():
				AportePatrimonial.objects.create(
					valor=movimentacao.valor,
					descricao='Aporte via Orçamento',
					id_transacao_origem=movimentacao.id,
				)

		if next_url:
			return redirect(next_url)

		return redirect('transacoes:livro_razao')

	context = _build_form_context()
	return render(request, 'transacoes/padrao/form_lancamento.html', context)


def partida_dupla(request):
	contas = ContaBancaria.objects.all().order_by('nome')
	plano_contas = PlanoConta.objects.all().order_by('codigo', 'nome')
	tags = Tag.objects.all().order_by('nome')
	hoje = timezone.localdate()

	if request.method == 'POST':
		try:
			data = json.loads(request.body)
			lancamentos = data.get('lancamentos', [])
			if not lancamentos or len(lancamentos) < 2:
				return JsonResponse({'error': 'Operação inválida.'}, status=400)
			with transaction.atomic():
				for l in lancamentos:
					mov = Movimentacao.objects.create(
						tipo=l.get('tipo'),
						valor=l.get('valor'),
						descricao=(l.get('descricao') or '').strip(),
						data_pagamento=l.get('data_pagamento'),
						data_vencimento=l.get('data_pagamento') or hoje,
						plano_conta_id=l.get('plano_conta_id'),
						conta_bancaria_id=l.get('conta_bancaria_id'),
						status=l.get('status', Movimentacao.Status.PENDENTE),
						formato_pagamento=l.get('formato_pagamento', 'PIX'),
					)
					# Salva tags se vierem
					tag_ids = l.get('tag_ids') or l.get('tags')
					if tag_ids:
						mov.tags.set(tag_ids)
			return JsonResponse({'ok': True})
		except Exception as e:
			return JsonResponse({'error': str(e)}, status=400)

	context = {
		'contas': contas,
		'plano_contas': plano_contas,
		'tags': tags,
		'hoje': hoje,
	}
	return render(request, 'transacoes/padrao/partida_dupla.html', context)


def lista_recorrentes(request):
	recorrentes = TransacaoRecorrente.objects.select_related(
		'plano_conta',
		'conta_bancaria',
	).prefetch_related(
		'tags',
	).order_by('plano_conta__codigo', 'plano_conta__nome', 'descricao')
	return render(request, 'transacoes/recorrentes/lista_recorrentes.html', {'recorrentes': recorrentes})


@require_http_methods(["POST"])
def excluir_recorrente(request, recorrente_id):
	recorrente = get_object_or_404(TransacaoRecorrente, pk=recorrente_id)
	next_url = request.POST.get('next') or request.GET.get('next')
	if next_url and not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
		next_url = None

	recorrente.delete()
	return redirect(next_url or 'transacoes:lista_recorrentes')


def nova_recorrente(request):
	contas = ContaBancaria.objects.all().order_by('nome')
	plano_contas = PlanoConta.objects.all().order_by('codigo', 'nome')
	tags = Tag.objects.all().order_by('nome')

	if request.method == 'POST':
		tag_ids = request.POST.getlist('tags')
		recorrente = TransacaoRecorrente.objects.create(
			descricao=request.POST.get('descricao', '').strip(),
			tipo=request.POST.get('tipo', 'Despesa'),
			plano_conta_id=request.POST.get('plano_conta_id'),
			conta_bancaria_id=request.POST.get('conta_bancaria_id'),
			formato_pagamento=request.POST.get('formato_pagamento', 'PIX'),
			dia_vencimento=request.POST.get('dia_vencimento') or 1,
			valor_base=request.POST.get('valor_base') or 0,
			tipo_valor=request.POST.get('tipo_valor', TransacaoRecorrente.TipoValor.EXATO),
			frequencia='Fixa',
			status_ativa=bool(request.POST.get('status_ativa')),
		)
		if tag_ids:
			recorrente.tags.set(tag_ids)
		return redirect('transacoes:lista_recorrentes')

	context = {
		'contas': contas,
		'plano_contas': plano_contas,
		'tags': tags,
	}
	return render(request, 'transacoes/recorrentes/nova_recorrente.html', context)


def lista_futuros(request):
	data_range = (request.GET.get('data_range') or '').strip()
	tag_id = (request.GET.get('tag_id') or '').strip()
	status_param = request.GET.get('status')
	status = LancamentoFuturo.Status.PENDENTE if status_param is None else (status_param or '').strip()
	valor_operador = (request.GET.get('valor_operador') or '').strip()
	valor_filtro = (request.GET.get('valor_filtro') or '').strip()

	inicio, fim = _parse_date_range(data_range)

	futuros = LancamentoFuturo.objects.select_related(
		'plano_conta',
		'conta_bancaria',
	).prefetch_related('tags').order_by('data_vencimento', 'descricao')

	if inicio:
		futuros = futuros.filter(data_vencimento__gte=inicio)
	if fim:
		futuros = futuros.filter(data_vencimento__lte=fim)
	if tag_id:
		futuros = futuros.filter(tags__id=tag_id)
	if status:
		futuros = futuros.filter(status=status)
	if valor_operador in {'gt', 'lt'} and valor_filtro:
		try:
			valor_normalizado = valor_filtro.replace('.', '').replace(',', '.')
			valor_decimal = Decimal(valor_normalizado)
			if valor_operador == 'gt':
				futuros = futuros.filter(valor__gt=valor_decimal)
			elif valor_operador == 'lt':
				futuros = futuros.filter(valor__lt=valor_decimal)
		except (InvalidOperation, ValueError):
			pass

	if tag_id:
		futuros = futuros.distinct()

	context = {
		'futuros': futuros,
		'tags': Tag.objects.all().order_by('nome'),
		'ciclo_ativo': Ciclo.objects.filter(status=Ciclo.Status.ABERTO).first(),
		'hoje': timezone.localdate(),
		'status_choices': LancamentoFuturo.Status.choices,
		'filtros': {
			'data_range': data_range,
			'tag_id': tag_id,
			'status': status,
			'valor_operador': valor_operador,
			'valor_filtro': valor_filtro,
		},
	}
	return render(request, 'transacoes/futuros/lista_futuros.html', context)


@require_http_methods(["POST"])
def excluir_futuro(request, futuro_id):
	futuro = get_object_or_404(LancamentoFuturo, pk=futuro_id)
	next_url = request.POST.get('next') or request.GET.get('next')
	if next_url and not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
		next_url = None

	futuro.delete()
	return redirect(next_url or 'transacoes:lista_futuros')


@require_http_methods(["POST"])
def adiantar_futuro(request, futuro_id):
	futuro = get_object_or_404(LancamentoFuturo, pk=futuro_id)
	next_url = request.POST.get('next') or request.GET.get('next')
	if next_url and not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
		next_url = None

	if futuro.status in {LancamentoFuturo.Status.INJETADO, LancamentoFuturo.Status.CANCELADO}:
		return redirect(next_url or 'transacoes:lista_futuros')

	modo_adiantamento = (request.POST.get('modo_adiantamento') or 'total').strip().lower()
	destino = (request.POST.get('destino') or 'avulso').strip().lower()
	data_pagamento = parse_date(request.POST.get('data_pagamento') or '') or timezone.localdate()

	valor_original = Decimal(futuro.valor or 0)
	if valor_original <= Decimal('0.00'):
		return redirect(next_url or 'transacoes:lista_futuros')

	if modo_adiantamento == 'parcial':
		try:
			valor_adiantado = Decimal((request.POST.get('valor_adiantamento') or '0').replace('.', '').replace(',', '.'))
		except (InvalidOperation, ValueError, TypeError):
			valor_adiantado = Decimal('0.00')
	else:
		valor_adiantado = valor_original

	if valor_adiantado <= Decimal('0.00') or valor_adiantado > valor_original:
		return redirect(next_url or 'transacoes:lista_futuros')

	ciclo_id = None
	status_movimentacao = Movimentacao.Status.VALIDADO
	if destino == 'ciclo':
		ciclo_ativo = Ciclo.objects.filter(status=Ciclo.Status.ABERTO).first()
		if not ciclo_ativo:
			return redirect(next_url or 'transacoes:lista_futuros')
		ciclo_id = ciclo_ativo.id
		status_movimentacao = Movimentacao.Status.PENDENTE

	with transaction.atomic():
		movimentacao = Movimentacao.objects.create(
			tipo=futuro.tipo,
			valor=valor_adiantado,
			descricao=futuro.descricao,
			data_pagamento=data_pagamento,
			data_vencimento=data_pagamento,
			conta_bancaria_id=futuro.conta_bancaria_id,
			conta_destino_id=futuro.conta_destino_id,
			plano_conta_id=futuro.plano_conta_id,
			status=status_movimentacao,
			formato_pagamento=futuro.formato_pagamento,
			frequencia=futuro.frequencia,
			ciclo_id=ciclo_id,
			cofre_id=futuro.cofre_id,
		)
		movimentacao.tags.set(futuro.tags.all())

		valor_restante = valor_original - valor_adiantado
		if valor_restante <= Decimal('0.00'):
			futuro.valor = Decimal('0.00')
			futuro.status = LancamentoFuturo.Status.INJETADO
		else:
			futuro.valor = valor_restante
			futuro.status = LancamentoFuturo.Status.PENDENTE
		futuro.save(update_fields=['valor', 'status', 'updated_at'])

	return redirect(next_url or 'transacoes:lista_futuros')


def novo_futuro(request):
	contas = ContaBancaria.objects.all().order_by('nome')
	plano_contas = PlanoConta.objects.all().order_by('codigo', 'nome')
	tags = Tag.objects.all().order_by('nome')

	if request.method == 'POST':
		tag_ids = request.POST.getlist('tags')
		descricao = request.POST.get('descricao', '').strip()
		tipo = request.POST.get('tipo', 'Despesa')
		plano_conta_id = request.POST.get('plano_conta_id')
		conta_bancaria_id = request.POST.get('conta_bancaria_id')
		formato_pagamento = request.POST.get('formato_pagamento', 'PIX')
		frequencia = request.POST.get('frequencia', 'Variavel')
		data_vencimento_str = request.POST.get('data_vencimento')
		valor_total_str = request.POST.get('valor') or '0'
		status = request.POST.get('status', LancamentoFuturo.Status.PENDENTE)
		total_parcelas = int(request.POST.get('total_parcelas') or 1)
		total_parcelas = max(total_parcelas, 1)
		comprovante_upload = request.FILES.get('comprovante')
		modo_lancamento = request.POST.get('modo_lancamento', 'unico')
		parcelas_valores_json = request.POST.get('parcelas_valores_json', '[]')
		parcelas_datas_json = request.POST.get('parcelas_datas_json', '[]')

		data_base = datetime.strptime(data_vencimento_str, '%Y-%m-%d').date()

		try:
			valor_total = Decimal(valor_total_str)
		except (InvalidOperation, TypeError):
			valor_total = Decimal('0')

		parcelas_custom = []
		if modo_lancamento == 'parcelado':
			try:
				valores_custom = json.loads(parcelas_valores_json)
				datas_custom = json.loads(parcelas_datas_json)
			except json.JSONDecodeError:
				valores_custom = []
				datas_custom = []

			if len(valores_custom) == total_parcelas and len(datas_custom) == total_parcelas:
				soma_parcelas = Decimal('0')
				for indice in range(total_parcelas):
					try:
						valor_parcela = Decimal(str(valores_custom[indice]))
						data_parcela = datetime.strptime(datas_custom[indice], '%Y-%m-%d').date()
					except (InvalidOperation, TypeError, ValueError):
						valor_parcela = Decimal('0')
						data_parcela = _add_months(data_base, indice)
					parcelas_custom.append((valor_parcela, data_parcela))
					soma_parcelas += valor_parcela

				if soma_parcelas.quantize(Decimal('0.01')) != valor_total.quantize(Decimal('0.01')):
					context = {
						'contas': contas,
						'plano_contas': plano_contas,
						'tags': tags,
						'hoje': timezone.localdate(),
						'erro_parcelas': 'A soma das parcelas deve ser igual ao valor total da compra.',
					}
					return render(request, 'transacoes/futuros/novo_futuro.html', context)

		primeiro = None
		for parcela in range(1, total_parcelas + 1):
			if parcelas_custom:
				valor_parcela, vencimento_parcela = parcelas_custom[parcela - 1]
			else:
				vencimento_parcela = _add_months(data_base, parcela - 1)
				valor_parcela = valor_total
			kwargs = {
				'descricao': descricao,
				'tipo': tipo,
				'plano_conta_id': plano_conta_id,
				'conta_bancaria_id': conta_bancaria_id,
				'formato_pagamento': formato_pagamento,
				'frequencia': frequencia,
				'data_vencimento': vencimento_parcela,
				'valor': valor_parcela,
				'status': status,
				'parcela_atual': parcela if total_parcelas > 1 else None,
				'total_parcelas': total_parcelas if total_parcelas > 1 else None,
			}

			if parcela == 1:
				if comprovante_upload:
					kwargs['comprovante'] = comprovante_upload
				futuro = LancamentoFuturo.objects.create(**kwargs)
				primeiro = futuro
			else:
				kwargs['lancamento_pai'] = primeiro
				if primeiro and primeiro.comprovante:
					kwargs['comprovante'] = primeiro.comprovante.name
				futuro = LancamentoFuturo.objects.create(**kwargs)

			if tag_ids:
				futuro.tags.set(tag_ids)
		return redirect('transacoes:lista_futuros')

	context = {
		'contas': contas,
		'plano_contas': plano_contas,
		'tags': tags,
		'hoje': timezone.localdate(),
	}
	return render(request, 'transacoes/futuros/novo_futuro.html', context)


@xframe_options_sameorigin
def painel_edicao(request, origem, registro_id):
	config = {
		'movimentacao': {
			'model': Movimentacao,
			'titulo': 'Editar Movimentacao',
			'lista_url': 'transacoes:livro_razao',
		},
		'futuro': {
			'model': LancamentoFuturo,
			'titulo': 'Editar Lancamento Futuro',
			'lista_url': 'transacoes:lista_futuros',
		},
		'recorrente': {
			'model': TransacaoRecorrente,
			'titulo': 'Editar Conta Fixa',
			'lista_url': 'transacoes:lista_recorrentes',
		},
	}

	if origem not in config:
		return redirect('core:home')

	next_url = request.POST.get('next') or request.GET.get('next')
	embed_mode = (request.GET.get('embed') == '1')
	if next_url and not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
		next_url = None

	default_redirect = config[origem]['lista_url']
	redirect_destino = next_url or default_redirect

	def _resposta_pos_salvar():
		if embed_mode:
			return render(
				request,
				'transacoes/padrao/painel_edicao_salvo_embed.html',
				{'redirect_url': redirect_destino},
			)
		return redirect(redirect_destino)

	model = config[origem]['model']
	registro = get_object_or_404(model, pk=registro_id)

	contas = ContaBancaria.objects.all().order_by('nome')
	plano_contas = PlanoConta.objects.all().order_by('codigo', 'nome')
	tags = Tag.objects.all().order_by('nome')
	erro_formulario = None

	if request.method == 'POST':
		if origem == 'movimentacao':
			tipo = request.POST.get('tipo')
			conta_origem_id = request.POST.get('conta_bancaria_id')
			conta_destino_id = request.POST.get('conta_destino_id') or None
			if tipo == TipoTransacao.TRANSFERENCIA and (not conta_destino_id or conta_destino_id == conta_origem_id):
				erro_formulario = 'Selecione uma conta destino diferente da origem para transferencia.'
			else:
				registro.tipo = tipo
				registro.valor = _to_decimal_or_zero(request.POST.get('valor'))
				registro.descricao = (request.POST.get('descricao') or '').strip()
				registro.formato_pagamento = request.POST.get('formato_pagamento') or registro.formato_pagamento
				registro.frequencia = request.POST.get('frequencia') or registro.frequencia
				registro.status = request.POST.get('status') or registro.status
				registro.data_pagamento = parse_date(request.POST.get('data_pagamento')) if request.POST.get('data_pagamento') else None
				registro.data_vencimento = parse_date(request.POST.get('data_vencimento')) or registro.data_vencimento
				registro.conta_bancaria_id = conta_origem_id or registro.conta_bancaria_id
				registro.conta_destino_id = conta_destino_id
				registro.plano_conta_id = request.POST.get('plano_conta_id') or registro.plano_conta_id
				if request.FILES.get('comprovante'):
					registro.comprovante = request.FILES.get('comprovante')
				registro.save()
				registro.tags.set(request.POST.getlist('tags'))
				return _resposta_pos_salvar()

		elif origem == 'futuro':
			registro.tipo = request.POST.get('tipo') or registro.tipo
			registro.valor = _to_decimal_or_zero(request.POST.get('valor'))
			registro.descricao = (request.POST.get('descricao') or '').strip()
			registro.formato_pagamento = request.POST.get('formato_pagamento') or registro.formato_pagamento
			registro.frequencia = request.POST.get('frequencia') or registro.frequencia
			registro.status = request.POST.get('status') or registro.status
			registro.data_vencimento = parse_date(request.POST.get('data_vencimento')) or registro.data_vencimento
			registro.conta_bancaria_id = request.POST.get('conta_bancaria_id') or registro.conta_bancaria_id
			registro.conta_destino_id = request.POST.get('conta_destino_id') or None
			registro.plano_conta_id = request.POST.get('plano_conta_id') or registro.plano_conta_id
			if request.FILES.get('comprovante'):
				registro.comprovante = request.FILES.get('comprovante')
			registro.save()
			registro.tags.set(request.POST.getlist('tags'))
			return _resposta_pos_salvar()

		elif origem == 'recorrente':
			tipo = request.POST.get('tipo') or registro.tipo
			conta_origem_id = request.POST.get('conta_bancaria_id')
			conta_destino_id = request.POST.get('conta_destino_id') or None
			if tipo == TipoTransacao.TRANSFERENCIA and (not conta_destino_id or conta_destino_id == conta_origem_id):
				erro_formulario = 'Selecione uma conta destino diferente da origem para transferencia.'
			else:
				registro.tipo = tipo
				registro.valor_base = _to_decimal_or_zero(request.POST.get('valor_base'))
				registro.descricao = (request.POST.get('descricao') or '').strip()
				registro.formato_pagamento = request.POST.get('formato_pagamento') or registro.formato_pagamento
				registro.dia_vencimento = int(request.POST.get('dia_vencimento') or registro.dia_vencimento or 1)
				registro.tipo_valor = request.POST.get('tipo_valor') or registro.tipo_valor
				registro.status_ativa = bool(request.POST.get('status_ativa'))
				registro.conta_bancaria_id = conta_origem_id or registro.conta_bancaria_id
				registro.conta_destino_id = conta_destino_id
				registro.plano_conta_id = request.POST.get('plano_conta_id') or registro.plano_conta_id
				registro.save()
				registro.tags.set(request.POST.getlist('tags'))
				return _resposta_pos_salvar()

	context = {
		'origem': origem,
		'registro': registro,
		'contas': contas,
		'plano_contas': plano_contas,
		'tags': tags,
		'titulo': config[origem]['titulo'],
		'lista_url': config[origem]['lista_url'],
		'retorno_url': redirect_destino,
		'next_url_form': next_url,
		'hide_sidebar': embed_mode,
		'erro_formulario': erro_formulario,
		'tipos_transacao': TipoTransacao.choices,
		'formatos_pagamento': Movimentacao._meta.get_field('formato_pagamento').choices,
		'frequencias': Movimentacao._meta.get_field('frequencia').choices,
		'mov_status': Movimentacao.Status.choices,
		'fut_status': LancamentoFuturo.Status.choices,
		'tipo_valor_choices': TransacaoRecorrente.TipoValor.choices,
	}
	return render(request, 'transacoes/padrao/painel_edicao.html', context)
