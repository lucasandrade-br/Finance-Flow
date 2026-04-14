from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
import re

from django.db import transaction
from django.db.models import Case, DecimalField, ExpressionWrapper, F, Q, Sum, Value, When
from django.db.models.functions import Coalesce, ExtractMonth, ExtractYear
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.dateparse import parse_date
from django.utils import timezone
from django.views.decorators.http import require_http_methods

from apps.contas.models import ContaBancaria
from apps.investimentos.models import AportePatrimonial, Ativo, MetaFinanceira, MetaParcelaMensal, Ordem, Rendimento
from apps.investimentos.services import calcular_rebalanceamento, processar_ordem, processar_rendimento, recalcular_posicao_ativo
from apps.transacoes.models import Movimentacao, TipoTransacao


STATUS_CONSOLIDADO = [Movimentacao.Status.EFETIVADO, Movimentacao.Status.VALIDADO]


def _month_start(dt):
	return dt.replace(day=1)


def _add_month(dt):
	if dt.month == 12:
		return dt.replace(year=dt.year + 1, month=1, day=1)
	return dt.replace(month=dt.month + 1, day=1)


def _iter_months(start_date, end_date):
	current = _month_start(start_date)
	end = _month_start(end_date)
	while current <= end:
		yield current
		current = _add_month(current)


def _to_decimal(value, default=Decimal('0')):
	try:
		return Decimal(str(value or '')).quantize(Decimal('0.01'))
	except (InvalidOperation, TypeError, ValueError):
		return default


def _to_decimal_raw(value, default=Decimal('0')):
	raw = str(value or '').strip()
	if not raw:
		return default
	raw = raw.replace('R$', '').replace(' ', '')
	if ',' in raw and '.' in raw:
		if raw.rfind(',') > raw.rfind('.'):
			raw = raw.replace('.', '').replace(',', '.')
		else:
			raw = raw.replace(',', '')
	elif ',' in raw:
		raw = raw.replace('.', '').replace(',', '.')
	try:
		return Decimal(raw)
	except (InvalidOperation, TypeError, ValueError):
		return default


def _parse_date_range(raw):
	if not raw:
		return (None, None)
	partes = re.findall(r'\d{4}-\d{2}-\d{2}', raw)
	if not partes:
		return (None, None)
	inicio = parse_date(partes[0])
	fim = parse_date(partes[-1]) if len(partes) > 1 else inicio
	if inicio and fim and inicio > fim:
		inicio, fim = fim, inicio
	return (inicio, fim)


def _parse_bool_cell(value):
	if isinstance(value, bool):
		return value
	texto = str(value or '').strip().lower()
	return texto in {'1', 'true', 'sim', 's', 'yes', 'y'}


def _resolve_or_create_ativo(ticker, nome_ativo, tipo_ativo):
	ticker = (ticker or '').strip().upper()
	nome_ativo = (nome_ativo or '').strip()
	tipo_ativo = (tipo_ativo or '').strip()

	if tipo_ativo not in [item[0] for item in Ativo.Tipo.choices]:
		tipo_ativo = Ativo.Tipo.ACAO

	if ticker:
		ativo = Ativo.objects.filter(ticker=ticker).first()
		if ativo:
			if nome_ativo and ativo.nome != nome_ativo:
				ativo.nome = nome_ativo
				ativo.save(update_fields=['nome', 'updated_at'])
			return ativo

		return Ativo.objects.create(
			nome=nome_ativo or ticker,
			ticker=ticker,
			tipo=tipo_ativo,
		)

	if not nome_ativo:
		raise ValueError('Informe ticker ou nome_ativo para identificar o ativo.')

	ativo = Ativo.objects.filter(nome__iexact=nome_ativo, tipo=tipo_ativo).first()
	if ativo:
		return ativo

	return Ativo.objects.create(
		nome=nome_ativo,
		tipo=tipo_ativo,
	)


def _distribuir_valor_mensal(valor_alvo, quantidade):
	if quantidade <= 0:
		return []
	total_centavos = int((valor_alvo * 100).to_integral_value())
	base = total_centavos // quantidade
	resto = total_centavos % quantidade
	valores = []
	for idx in range(quantidade):
		centavos = base + (1 if idx < resto else 0)
		valores.append((Decimal(centavos) / Decimal('100')).quantize(Decimal('0.01')))
	return valores


def _regerar_parcelas_mensais(meta):
	competencias = list(_iter_months(meta.data_inicio, meta.data_fim))
	if not competencias:
		return False

	valores = _distribuir_valor_mensal(meta.valor_alvo, len(competencias))
	parcelas = []
	for idx, competencia in enumerate(competencias, start=1):
		parcelas.append(
			MetaParcelaMensal(
				meta=meta,
				competencia=competencia,
				valor_planejado=valores[idx - 1],
				ordem_mes=idx,
			)
		)

	meta.parcelas_mensais.all().delete()
	MetaParcelaMensal.objects.bulk_create(parcelas)
	return True


def _saldo_atual_por_contas(conta_ids):
	if not conta_ids:
		return {}

	agregados = (
		Movimentacao.objects.filter(conta_bancaria_id__in=conta_ids, status__in=STATUS_CONSOLIDADO)
		.values('conta_bancaria_id')
		.annotate(
			total_creditos=Coalesce(
				Sum(
					Case(
						When(tipo__in=[TipoTransacao.RECEITA, TipoTransacao.TRANSFERENCIA_ENTRADA], then=F('valor')),
						default=Value(0),
						output_field=DecimalField(max_digits=14, decimal_places=2),
					)
				),
				Value(Decimal('0.00')),
			),
			total_debitos=Coalesce(
				Sum(
					Case(
						When(
							tipo__in=[
								TipoTransacao.DESPESA,
								TipoTransacao.INVESTIMENTO,
								TipoTransacao.TRANSFERENCIA_SAIDA,
							],
							then=F('valor'),
						),
						default=Value(0),
						output_field=DecimalField(max_digits=14, decimal_places=2),
					)
				),
				Value(Decimal('0.00')),
			),
		)
	)

	mapa = {conta_id: Decimal('0.00') for conta_id in conta_ids}
	for item in agregados:
		conta_id = item['conta_bancaria_id']
		mapa[conta_id] = (item['total_creditos'] - item['total_debitos']).quantize(Decimal('0.01'))
	return mapa


def _realizado_mensal_meta(meta):
	base = (
		Movimentacao.objects.filter(
			conta_bancaria_id=meta.conta_bancaria_id,
			tipo__in=[TipoTransacao.TRANSFERENCIA_ENTRADA, TipoTransacao.RECEITA],
			status__in=STATUS_CONSOLIDADO,
		)
		.annotate(data_referencia=Coalesce('data_pagamento', 'data_vencimento'))
		.filter(data_referencia__range=(meta.data_inicio, meta.data_fim))
		.annotate(ano=ExtractYear('data_referencia'), mes=ExtractMonth('data_referencia'))
		.values('ano', 'mes')
		.annotate(total=Coalesce(Sum('valor'), Value(Decimal('0.00'))))
	)
	resultado = {}
	for item in base:
		competencia = date(item['ano'], item['mes'], 1)
		resultado[competencia] = item['total'].quantize(Decimal('0.01'))
	return resultado


@require_http_methods(["GET"])
def lista_metas(request):
	metas = list(MetaFinanceira.objects.select_related('conta_bancaria').order_by('-created_at'))
	conta_ids = [m.conta_bancaria_id for m in metas]
	saldo_movimentos_map = _saldo_atual_por_contas(conta_ids)

	for meta in metas:
		saldo_mov = saldo_movimentos_map.get(meta.conta_bancaria_id, Decimal('0.00'))
		saldo_atual = (meta.conta_bancaria.saldo_inicial + saldo_mov).quantize(Decimal('0.01'))
		meta.saldo_atual = saldo_atual

		realizado_total = sum(_realizado_mensal_meta(meta).values(), Decimal('0.00')).quantize(Decimal('0.01'))
		percentual = Decimal('0.00')
		if meta.valor_alvo and meta.valor_alvo > 0:
			percentual = ((realizado_total / meta.valor_alvo) * Decimal('100')).quantize(Decimal('0.01'))
		meta.realizado_total = realizado_total
		meta.percentual_alcancado = percentual

	context = {
		'metas': metas,
	}
	return render(request, 'investimentos/metas/lista.html', context)


@require_http_methods(["GET", "POST"])
def nova_meta(request):
	contas = ContaBancaria.objects.order_by('nome')
	erro_formulario = None

	if request.method == 'POST':
		nome = (request.POST.get('nome') or '').strip()
		conta_id = request.POST.get('conta_bancaria_id')
		valor_alvo = _to_decimal(request.POST.get('valor_alvo'))
		data_inicio_raw = request.POST.get('data_inicio')
		data_fim_raw = request.POST.get('data_fim')

		try:
			data_inicio = date.fromisoformat(data_inicio_raw)
			data_fim = date.fromisoformat(data_fim_raw)
		except (TypeError, ValueError):
			data_inicio = None
			data_fim = None

		if not nome:
			erro_formulario = 'Informe um nome para a meta.'
		elif not conta_id:
			erro_formulario = 'Selecione uma conta bancária.'
		elif valor_alvo <= 0:
			erro_formulario = 'O valor alvo deve ser maior que zero.'
		elif not data_inicio or not data_fim:
			erro_formulario = 'Informe um período válido para a meta.'
		elif data_fim < data_inicio:
			erro_formulario = 'A data final deve ser igual ou posterior à data inicial.'
		else:
			competencias = list(_iter_months(data_inicio, data_fim))
			if not competencias:
				erro_formulario = 'Não foi possível gerar competências mensais para o período informado.'
			else:
				with transaction.atomic():
					meta = MetaFinanceira.objects.create(
						nome=nome,
						conta_bancaria_id=conta_id,
						valor_alvo=valor_alvo,
						data_inicio=data_inicio,
						data_fim=data_fim,
						status=MetaFinanceira.Status.ATIVA,
					)
					_regerar_parcelas_mensais(meta)
				return redirect('investimentos:roteiro_meta', meta_id=meta.id)

	context = {
		'contas': contas,
		'erro_formulario': erro_formulario,
		'hoje': date.today(),
	}
	return render(request, 'investimentos/metas/nova.html', context)


@require_http_methods(["GET", "POST"])
def roteiro_meta(request, meta_id):
	meta = get_object_or_404(MetaFinanceira.objects.select_related('conta_bancaria'), pk=meta_id)
	parcelas = list(meta.parcelas_mensais.order_by('competencia'))
	erro_formulario = None

	if request.method == 'POST':
		soma = Decimal('0.00')
		for parcela in parcelas:
			valor = _to_decimal(request.POST.get(f'valor_planejado_{parcela.id}'))
			if valor < 0:
				erro_formulario = 'Os valores mensais não podem ser negativos.'
				break
			parcela.valor_planejado = valor
			soma += valor

		acao = request.POST.get('acao') or 'rascunho'
		if not erro_formulario and acao == 'ativar' and soma.quantize(Decimal('0.01')) != meta.valor_alvo.quantize(Decimal('0.01')):
			erro_formulario = 'A soma das parcelas deve ser igual ao valor alvo para ativar a meta.'

		if not erro_formulario:
			MetaParcelaMensal.objects.bulk_update(parcelas, ['valor_planejado', 'updated_at'])
			if acao == 'ativar':
				meta.status = MetaFinanceira.Status.ATIVA
			else:
				meta.status = MetaFinanceira.Status.RASCUNHO
			meta.save(update_fields=['status', 'updated_at'])
			return redirect('investimentos:roteiro_meta', meta_id=meta.id)

	realizado_map = _realizado_mensal_meta(meta)
	linhas = []
	total_planejado = Decimal('0.00')
	total_realizado = Decimal('0.00')
	for parcela in parcelas:
		realizado = realizado_map.get(parcela.competencia, Decimal('0.00'))
		total_planejado += parcela.valor_planejado
		total_realizado += realizado
		linhas.append({
			'parcela': parcela,
			'realizado': realizado,
			'diferenca': (realizado - parcela.valor_planejado).quantize(Decimal('0.01')),
		})

	context = {
		'meta': meta,
		'linhas': linhas,
		'total_planejado': total_planejado.quantize(Decimal('0.01')),
		'total_realizado': total_realizado.quantize(Decimal('0.01')),
		'contas': ContaBancaria.objects.order_by('nome'),
		'erro_formulario': erro_formulario,
	}
	return render(request, 'investimentos/metas/roteiro.html', context)


@require_http_methods(["POST"])
def editar_meta(request, meta_id):
	meta = get_object_or_404(MetaFinanceira, pk=meta_id)

	nome = (request.POST.get('nome') or '').strip()
	conta_id = request.POST.get('conta_bancaria_id')
	valor_alvo = _to_decimal(request.POST.get('valor_alvo'))
	data_inicio_raw = request.POST.get('data_inicio')
	data_fim_raw = request.POST.get('data_fim')

	try:
		data_inicio = date.fromisoformat(data_inicio_raw)
		data_fim = date.fromisoformat(data_fim_raw)
	except (TypeError, ValueError):
		return redirect('investimentos:roteiro_meta', meta_id=meta.id)

	if not nome or not conta_id or valor_alvo <= 0 or data_fim < data_inicio:
		return redirect('investimentos:roteiro_meta', meta_id=meta.id)

	with transaction.atomic():
		meta.nome = nome
		meta.conta_bancaria_id = conta_id
		meta.valor_alvo = valor_alvo
		meta.data_inicio = data_inicio
		meta.data_fim = data_fim
		meta.save(update_fields=['nome', 'conta_bancaria', 'valor_alvo', 'data_inicio', 'data_fim', 'updated_at'])

		if not _regerar_parcelas_mensais(meta):
			return redirect('investimentos:roteiro_meta', meta_id=meta.id)

	return redirect('investimentos:roteiro_meta', meta_id=meta.id)


@require_http_methods(["POST"])
def excluir_meta(request, meta_id):
	meta = get_object_or_404(MetaFinanceira, pk=meta_id)
	meta.delete()
	return redirect('investimentos:lista_metas')


@require_http_methods(["GET"])
def painel_meta(request, meta_id):
	meta = get_object_or_404(MetaFinanceira.objects.select_related('conta_bancaria'), pk=meta_id)
	parcelas = list(meta.parcelas_mensais.order_by('competencia'))
	realizado_map = _realizado_mensal_meta(meta)

	labels = []
	planejado_series = []
	realizado_series = []
	acumulado_planejado_series = []
	acumulado_realizado_series = []

	acumulado_planejado = Decimal('0.00')
	acumulado_realizado = Decimal('0.00')
	for parcela in parcelas:
		realizado = realizado_map.get(parcela.competencia, Decimal('0.00'))
		acumulado_planejado += parcela.valor_planejado
		acumulado_realizado += realizado

		labels.append(parcela.competencia.strftime('%m/%Y'))
		planejado_series.append(float(parcela.valor_planejado))
		realizado_series.append(float(realizado))
		acumulado_planejado_series.append(float(acumulado_planejado))
		acumulado_realizado_series.append(float(acumulado_realizado))

	percentual = Decimal('0.00')
	if meta.valor_alvo > 0:
		percentual = ((acumulado_realizado / meta.valor_alvo) * Decimal('100')).quantize(Decimal('0.01'))

	context = {
		'meta': meta,
		'valor_realizado': acumulado_realizado.quantize(Decimal('0.01')),
		'valor_faltante': (meta.valor_alvo - acumulado_realizado).quantize(Decimal('0.01')),
		'percentual_alcancado': percentual,
		'labels': labels,
		'planejado_series': planejado_series,
		'realizado_series': realizado_series,
		'acumulado_planejado_series': acumulado_planejado_series,
		'acumulado_realizado_series': acumulado_realizado_series,
	}
	return render(request, 'investimentos/metas/painel.html', context)


@require_http_methods(["GET", "POST"])
def painel_investimentos(request):
	ativos = list(Ativo.objects.order_by('ticker', 'nome'))
	patrimonio_total = Decimal('0.00')
	saldo_disponivel_rebalance = AportePatrimonial.saldo_disponivel()
	aportes_recentes = AportePatrimonial.objects.order_by('-data')[:12]

	for ativo in ativos:
		valor_atual = (Decimal(ativo.quantidade_atual or 0) * Decimal(ativo.preco_medio or 0)).quantize(Decimal('0.01'))
		ativo.valor_atual = valor_atual
		patrimonio_total += valor_atual

	for ativo in ativos:
		if patrimonio_total > 0:
			ativo.percentual_atual = ((ativo.valor_atual / patrimonio_total) * Decimal('100')).quantize(Decimal('0.01'))
		else:
			ativo.percentual_atual = Decimal('0.00')

	ativos.sort(key=lambda ativo: ativo.percentual_atual, reverse=True)

	rebalanceamento = None
	valor_aporte = saldo_disponivel_rebalance
	erro_formulario = None

	if request.method == 'POST':
		valor_aporte = _to_decimal(request.POST.get('valor_aporte'), default=saldo_disponivel_rebalance)
		if valor_aporte <= 0:
			erro_formulario = 'Informe um valor de aporte maior que zero para analisar o rebalanceamento.'
		else:
			rebalanceamento = calcular_rebalanceamento(valor_aporte)

	context = {
		'ativos': ativos,
		'patrimonio_total': patrimonio_total.quantize(Decimal('0.01')),
		'valor_aporte': valor_aporte,
		'saldo_disponivel_rebalance': saldo_disponivel_rebalance,
		'aportes_recentes': aportes_recentes,
		'rebalanceamento': rebalanceamento,
		'erro_formulario': erro_formulario,
	}
	return render(request, 'investimentos/painel.html', context)


@require_http_methods(["GET"])
def relatorios_investimentos(request, ano=None):
	hoje = timezone.localdate()
	ano_get = (request.GET.get('ano') or '').strip()
	try:
		ano_alvo = int(ano_get) if ano_get else (ano or hoje.year)
	except ValueError:
		ano_alvo = ano or hoje.year

	comparar_anos_ativo = (request.GET.get('comparar_anos') == '1')
	agrupamento = (request.GET.get('agrupamento') or 'anos').strip().lower()
	if agrupamento not in ['anos', 'meses']:
		agrupamento = 'anos'

	anos_comparar_raw = request.GET.getlist('anos_comparar')
	anos_comparar = []
	for ano_item in anos_comparar_raw:
		try:
			anos_comparar.append(int(str(ano_item).strip()))
		except (TypeError, ValueError):
			continue
	anos_comparar = sorted(set(anos_comparar))
	if not comparar_anos_ativo:
		anos_comparar = [ano_alvo]
	elif not anos_comparar:
		anos_comparar = [ano_alvo]
	modo_comparativo = comparar_anos_ativo

	ativo_id = (request.GET.get('ativo_id') or '').strip()
	tipo_ativo = (request.GET.get('tipo_ativo') or '').strip()
	setor = (request.GET.get('setor') or '').strip()
	tipos_validos = [item[0] for item in Ativo.Tipo.choices]

	ordens_base = Ordem.objects.select_related('ativo')
	rendimentos_base = Rendimento.objects.select_related('ativo')

	if ativo_id and ativo_id.isdigit():
		ordens_base = ordens_base.filter(ativo_id=ativo_id)
		rendimentos_base = rendimentos_base.filter(ativo_id=ativo_id)

	if tipo_ativo in tipos_validos:
		ordens_base = ordens_base.filter(ativo__tipo=tipo_ativo)
		rendimentos_base = rendimentos_base.filter(ativo__tipo=tipo_ativo)

	if setor:
		ordens_base = ordens_base.filter(ativo__setor__iexact=setor)
		rendimentos_base = rendimentos_base.filter(ativo__setor__iexact=setor)

	valor_ordem_expr = ExpressionWrapper(
		F('quantidade') * F('preco') + F('taxas'),
		output_field=DecimalField(max_digits=24, decimal_places=8),
	)

	ordens_compra_mes = (
		ordens_base.filter(tipo=Ordem.TipoOrdem.COMPRA, data__year=ano_alvo)
		.annotate(mes=ExtractMonth('data'))
		.values('mes')
		.annotate(total=Coalesce(Sum(valor_ordem_expr), Value(Decimal('0.00'))))
	)
	ordens_venda_mes = (
		ordens_base.filter(tipo=Ordem.TipoOrdem.VENDA, data__year=ano_alvo)
		.annotate(mes=ExtractMonth('data'))
		.values('mes')
		.annotate(total=Coalesce(Sum(valor_ordem_expr), Value(Decimal('0.00'))))
	)
	rendimentos_mes = (
		rendimentos_base.filter(data__year=ano_alvo)
		.annotate(mes=ExtractMonth('data'))
		.values('mes')
		.annotate(total=Coalesce(Sum('valor'), Value(Decimal('0.00'))))
	)

	compras_map = {item['mes']: Decimal(item['total'] or 0).quantize(Decimal('0.01')) for item in ordens_compra_mes}
	vendas_map = {item['mes']: Decimal(item['total'] or 0).quantize(Decimal('0.01')) for item in ordens_venda_mes}
	rendas_map = {item['mes']: Decimal(item['total'] or 0).quantize(Decimal('0.01')) for item in rendimentos_mes}

	meses_labels = [
		'Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez'
	]
	matrix = []
	total_capital_mes = Decimal('0.00')
	total_renda_mes = Decimal('0.00')
	compras_series = []
	rendas_series = []
	patrimonio_series = []

	for mes in range(1, 13):
		compras_mes = compras_map.get(mes, Decimal('0.00')).quantize(Decimal('0.01'))
		vendas_mes = vendas_map.get(mes, Decimal('0.00')).quantize(Decimal('0.01'))
		capital_mes = (compras_mes - vendas_mes).quantize(Decimal('0.01'))
		total_capital_mes += capital_mes

		ultimo_dia_mes = date(ano_alvo, mes, 1)
		if mes == 12:
			ultimo_dia_mes = ultimo_dia_mes.replace(day=31)
		else:
			ultimo_dia_mes = (date(ano_alvo, mes + 1, 1) - timedelta(days=1))

		total_compras_acum = (
			ordens_base.filter(tipo=Ordem.TipoOrdem.COMPRA, data__lte=ultimo_dia_mes)
			.aggregate(total=Coalesce(Sum(valor_ordem_expr), Value(Decimal('0.00'))))
			['total']
		)
		total_vendas_acum = (
			ordens_base.filter(tipo=Ordem.TipoOrdem.VENDA, data__lte=ultimo_dia_mes)
			.aggregate(total=Coalesce(Sum(valor_ordem_expr), Value(Decimal('0.00'))))
			['total']
		)
		capital_acumulado = (Decimal(total_compras_acum or 0) - Decimal(total_vendas_acum or 0)).quantize(Decimal('0.01'))

		renda_mes = rendas_map.get(mes, Decimal('0.00')).quantize(Decimal('0.01'))
		total_renda_mes += renda_mes

		yield_mes = Decimal('0.00')
		if capital_acumulado > 0:
			yield_mes = ((renda_mes / capital_acumulado) * Decimal('100')).quantize(Decimal('0.01'))

		matrix.append(
			{
				'mes': mes,
				'mes_label': meses_labels[mes - 1],
				'compras_mes': compras_mes,
				'vendas_mes': vendas_mes,
				'capital_mes': capital_mes,
				'capital_acumulado': capital_acumulado,
				'renda_mes': renda_mes,
				'yield_mes': yield_mes,
			}
		)

		compras_series.append(float(compras_mes))
		rendas_series.append(float(renda_mes))
		patrimonio_series.append(float(capital_acumulado))

	capital_final_ano = matrix[-1]['capital_acumulado'] if matrix else Decimal('0.00')
	yield_anual = Decimal('0.00')
	if capital_final_ano > 0:
		yield_anual = ((total_renda_mes / capital_final_ano) * Decimal('100')).quantize(Decimal('0.01'))

	anos_disponiveis = sorted(
		set(
			list(ordens_base.values_list('data__year', flat=True))
			+ list(rendimentos_base.values_list('data__year', flat=True))
		),
		reverse=True,
	)
	if ano_alvo not in anos_disponiveis:
		anos_disponiveis = sorted(set(anos_disponiveis + [ano_alvo]), reverse=True)

	setores_disponiveis = (
		Ativo.objects.exclude(setor='')
		.values_list('setor', flat=True)
		.distinct()
		.order_by('setor')
	)
	ativos_disponiveis = Ativo.objects.order_by('ticker', 'nome')

	def _resumo_por_ano(ano_ref):
		total_compras_ano = (
			ordens_base.filter(tipo=Ordem.TipoOrdem.COMPRA, data__year=ano_ref)
			.aggregate(total=Coalesce(Sum(valor_ordem_expr), Value(Decimal('0.00'))))
			['total']
		)
		total_vendas_ano = (
			ordens_base.filter(tipo=Ordem.TipoOrdem.VENDA, data__year=ano_ref)
			.aggregate(total=Coalesce(Sum(valor_ordem_expr), Value(Decimal('0.00'))))
			['total']
		)
		total_rendas_ano = (
			rendimentos_base.filter(data__year=ano_ref)
			.aggregate(total=Coalesce(Sum('valor'), Value(Decimal('0.00'))))
			['total']
		)

		capital_ano = (Decimal(total_compras_ano or 0) - Decimal(total_vendas_ano or 0)).quantize(Decimal('0.01'))
		renda_ano = Decimal(total_rendas_ano or 0).quantize(Decimal('0.01'))

		ultimo_dia_ano = date(ano_ref, 12, 31)
		total_compras_acum = (
			ordens_base.filter(tipo=Ordem.TipoOrdem.COMPRA, data__lte=ultimo_dia_ano)
			.aggregate(total=Coalesce(Sum(valor_ordem_expr), Value(Decimal('0.00'))))
			['total']
		)
		total_vendas_acum = (
			ordens_base.filter(tipo=Ordem.TipoOrdem.VENDA, data__lte=ultimo_dia_ano)
			.aggregate(total=Coalesce(Sum(valor_ordem_expr), Value(Decimal('0.00'))))
			['total']
		)
		capital_final = (Decimal(total_compras_acum or 0) - Decimal(total_vendas_acum or 0)).quantize(Decimal('0.01'))

		yield_anual_item = Decimal('0.00')
		if capital_final > 0:
			yield_anual_item = ((renda_ano / capital_final) * Decimal('100')).quantize(Decimal('0.01'))

		return {
			'ano': ano_ref,
			'compras_ano': Decimal(total_compras_ano or 0).quantize(Decimal('0.01')),
			'vendas_ano': Decimal(total_vendas_ano or 0).quantize(Decimal('0.01')),
			'capital_ano': capital_ano,
			'renda_ano': renda_ano,
			'capital_final': capital_final,
			'yield_anual': yield_anual_item,
		}

	resumos_comparativos = [_resumo_por_ano(ano_ref) for ano_ref in anos_comparar]
	grafico_anual = {
		'labels': [str(item['ano']) for item in resumos_comparativos],
		'compras': [float(item['compras_ano']) for item in resumos_comparativos],
		'rendas': [float(item['renda_ano']) for item in resumos_comparativos],
		'patrimonio': [float(item['capital_final']) for item in resumos_comparativos],
	}

	if agrupamento == 'meses' and anos_comparar:
		labels_meses = []
		compras_meses = []
		rendas_meses = []
		patrimonio_meses = []

		for ano_ref in anos_comparar:
			compras_mes_ano = (
				ordens_base.filter(tipo=Ordem.TipoOrdem.COMPRA, data__year=ano_ref)
				.annotate(mes=ExtractMonth('data'))
				.values('mes')
				.annotate(total=Coalesce(Sum(valor_ordem_expr), Value(Decimal('0.00'))))
			)
			vendas_mes_ano = (
				ordens_base.filter(tipo=Ordem.TipoOrdem.VENDA, data__year=ano_ref)
				.annotate(mes=ExtractMonth('data'))
				.values('mes')
				.annotate(total=Coalesce(Sum(valor_ordem_expr), Value(Decimal('0.00'))))
			)
			rendas_mes_ano = (
				rendimentos_base.filter(data__year=ano_ref)
				.annotate(mes=ExtractMonth('data'))
				.values('mes')
				.annotate(total=Coalesce(Sum('valor'), Value(Decimal('0.00'))))
			)

			compras_aux = {item['mes']: Decimal(item['total'] or 0).quantize(Decimal('0.01')) for item in compras_mes_ano}
			vendas_aux = {item['mes']: Decimal(item['total'] or 0).quantize(Decimal('0.01')) for item in vendas_mes_ano}
			rendas_aux = {item['mes']: Decimal(item['total'] or 0).quantize(Decimal('0.01')) for item in rendas_mes_ano}

			for mes in range(1, 13):
				compras_mes_item = compras_aux.get(mes, Decimal('0.00'))
				vendas_mes_item = vendas_aux.get(mes, Decimal('0.00'))
				renda_mes_item = rendas_aux.get(mes, Decimal('0.00'))

				if mes == 12:
					ultimo_dia_mes = date(ano_ref, 12, 31)
				else:
					ultimo_dia_mes = date(ano_ref, mes + 1, 1) - timedelta(days=1)

				total_compras_acum = (
					ordens_base.filter(tipo=Ordem.TipoOrdem.COMPRA, data__lte=ultimo_dia_mes)
					.aggregate(total=Coalesce(Sum(valor_ordem_expr), Value(Decimal('0.00'))))
					['total']
				)
				total_vendas_acum = (
					ordens_base.filter(tipo=Ordem.TipoOrdem.VENDA, data__lte=ultimo_dia_mes)
					.aggregate(total=Coalesce(Sum(valor_ordem_expr), Value(Decimal('0.00'))))
					['total']
				)
				capital_acumulado_global = (
					Decimal(total_compras_acum or 0) - Decimal(total_vendas_acum or 0)
				).quantize(Decimal('0.01'))

				labels_meses.append(f"{meses_labels[mes - 1]}/{ano_ref}")
				compras_meses.append(float(compras_mes_item))
				rendas_meses.append(float(renda_mes_item))
				patrimonio_meses.append(float(capital_acumulado_global))

		grafico_comparativo = {
			'labels': labels_meses,
			'compras': compras_meses,
			'rendas': rendas_meses,
			'patrimonio': patrimonio_meses,
		}
	else:
		grafico_comparativo = grafico_anual

	context = {
		'ano_alvo': ano_alvo,
		'modo_comparativo': modo_comparativo,
		'comparar_anos_ativo': comparar_anos_ativo,
		'agrupamento': agrupamento,
		'anos_comparar': [str(item) for item in anos_comparar],
		'anos_disponiveis': anos_disponiveis,
		'matriz_analitica': matrix,
		'ativos_disponiveis': ativos_disponiveis,
		'setores_disponiveis': setores_disponiveis,
		'tipos_ativo': Ativo.Tipo.choices,
		'filtros': {
			'ativo_id': ativo_id,
			'tipo_ativo': tipo_ativo,
			'setor': setor,
		},
		'resumos_comparativos': resumos_comparativos,
		'grafico_comparativo': grafico_comparativo,
		'grafico': {
			'labels': meses_labels,
			'patrimonio': patrimonio_series,
			'compras': compras_series,
			'rendas': rendas_series,
		},
		'grafico_anual': grafico_anual,
		'resumo_anual': {
			'capital_ano': total_capital_mes.quantize(Decimal('0.01')),
			'renda_ano': total_renda_mes.quantize(Decimal('0.01')),
			'capital_final': capital_final_ano.quantize(Decimal('0.01')),
			'yield_anual': yield_anual,
		},
	}
	return render(request, 'investimentos/relatorios.html', context)


@require_http_methods(["GET", "POST"])
def nova_ordem(request):
	ativos = Ativo.objects.order_by('ticker', 'nome')
	erro_formulario = None

	if request.method == 'POST':
		ativo_id = request.POST.get('ativo_id')
		tipo = request.POST.get('tipo')
		quantidade = _to_decimal(request.POST.get('quantidade'))
		preco = _to_decimal(request.POST.get('preco'))
		taxas = _to_decimal(request.POST.get('taxas'))
		data_raw = request.POST.get('data')

		try:
			data_ordem = date.fromisoformat(data_raw)
		except (TypeError, ValueError):
			data_ordem = None

		if not ativo_id:
			erro_formulario = 'Selecione um ativo para registrar a ordem.'
		elif tipo not in [Ordem.TipoOrdem.COMPRA, Ordem.TipoOrdem.VENDA]:
			erro_formulario = 'Tipo de ordem inválido.'
		elif quantidade <= 0:
			erro_formulario = 'A quantidade deve ser maior que zero.'
		elif preco <= 0:
			erro_formulario = 'O preço deve ser maior que zero.'
		elif not data_ordem:
			erro_formulario = 'Informe uma data válida para a ordem.'
		else:
			try:
				with transaction.atomic():
					ordem = Ordem.objects.create(
						ativo_id=ativo_id,
						tipo=tipo,
						quantidade=quantidade,
						preco=preco,
						taxas=taxas,
						data=data_ordem,
					)
					processar_ordem(ordem)
			except ValueError as exc:
				erro_formulario = str(exc)
			else:
				return redirect('investimentos:painel_investimentos')

	context = {
		'ativos': ativos,
		'erro_formulario': erro_formulario,
		'hoje': date.today(),
	}
	return render(request, 'investimentos/ordem_form.html', context)


@require_http_methods(["GET", "POST"])
def novo_rendimento(request):
	ativos = Ativo.objects.order_by('ticker', 'nome')
	erro_formulario = None

	if request.method == 'POST':
		ativo_id = request.POST.get('ativo_id')
		descricao = (request.POST.get('descricao') or '').strip()
		resgatar_para_orcamento = (request.POST.get('resgatar_para_orcamento') == 'on')
		valores_raw = request.POST.getlist('valor_item')
		datas_raw = request.POST.getlist('data_item')

		itens = []
		total_linhas = max(len(valores_raw), len(datas_raw))

		for idx in range(total_linhas):
			valor_raw = (valores_raw[idx] if idx < len(valores_raw) else '').strip()
			data_raw = (datas_raw[idx] if idx < len(datas_raw) else '').strip()

			if not valor_raw and not data_raw:
				continue

			valor = _to_decimal(valor_raw)
			if valor <= 0:
				erro_formulario = f'O valor da linha {idx + 1} deve ser maior que zero.'
				break

			try:
				data_rendimento = date.fromisoformat(data_raw)
			except (TypeError, ValueError):
				erro_formulario = f'Informe uma data válida na linha {idx + 1}.'
				break

			itens.append({'valor': valor, 'data': data_rendimento})

		if not ativo_id:
			erro_formulario = 'Selecione o ativo do rendimento.'
		elif not itens and not erro_formulario:
			erro_formulario = 'Adicione pelo menos um rendimento com valor e data.'
		else:
			try:
				with transaction.atomic():
					for item in itens:
						rendimento = Rendimento.objects.create(
							ativo_id=ativo_id,
							valor=item['valor'],
							data=item['data'],
							descricao=descricao,
							resgatar_para_orcamento=resgatar_para_orcamento,
						)
						processar_rendimento(rendimento)
			except ValueError as exc:
				erro_formulario = str(exc)
			else:
				return redirect('investimentos:painel_investimentos')

	context = {
		'ativos': ativos,
		'erro_formulario': erro_formulario,
		'hoje': date.today(),
	}
	return render(request, 'investimentos/rendimento_form.html', context)


@require_http_methods(["GET", "POST"])
def gestao_ativos(request):
	erro_formulario = None
	sucesso_formulario = None
	acao = request.POST.get('acao')
	tipos_validos = [item[0] for item in Ativo.Tipo.choices]

	if request.method == 'POST':
		if acao in ['criar', 'editar']:
			ativo_id = request.POST.get('ativo_id')
			nome = (request.POST.get('nome') or '').strip()
			ticker = (request.POST.get('ticker') or '').strip().upper()
			setor = (request.POST.get('setor') or '').strip()
			tipo = request.POST.get('tipo')
			quantidade_atual = _to_decimal_raw(request.POST.get('quantidade_atual')).quantize(Decimal('0.00000001'))
			preco_medio = _to_decimal_raw(request.POST.get('preco_medio')).quantize(Decimal('0.00000001'))
			percentual_alvo = _to_decimal_raw(request.POST.get('percentual_alvo')).quantize(Decimal('0.01'))

			if not nome:
				erro_formulario = 'Informe o nome do ativo.'
			elif tipo not in tipos_validos:
				erro_formulario = 'Tipo de ativo inválido.'
			elif len(setor) > 80:
				erro_formulario = 'O setor deve ter no máximo 80 caracteres.'
			elif quantidade_atual < 0 or preco_medio < 0:
				erro_formulario = 'Quantidade e preço médio não podem ser negativos.'
			elif percentual_alvo < 0 or percentual_alvo > 100:
				erro_formulario = 'O percentual alvo deve ficar entre 0 e 100.'
			else:
				filtro_ticker = Ativo.objects.filter(ticker=ticker) if ticker else Ativo.objects.none()
				if acao == 'editar' and ativo_id:
					filtro_ticker = filtro_ticker.exclude(pk=ativo_id)
				if ticker and filtro_ticker.exists():
					erro_formulario = 'Já existe um ativo com este ticker.'
				else:
					if acao == 'criar':
						Ativo.objects.create(
							nome=nome,
							ticker=ticker or None,
							setor=setor,
							tipo=tipo,
							quantidade_atual=quantidade_atual,
							preco_medio=preco_medio,
							percentual_alvo=percentual_alvo,
						)
						sucesso_formulario = 'Ativo cadastrado com sucesso.'
					else:
						ativo = get_object_or_404(Ativo, pk=ativo_id)
						ativo.nome = nome
						ativo.ticker = ticker or None
						ativo.setor = setor
						ativo.tipo = tipo
						ativo.quantidade_atual = quantidade_atual
						ativo.preco_medio = preco_medio
						ativo.percentual_alvo = percentual_alvo
						ativo.save(update_fields=['nome', 'ticker', 'setor', 'tipo', 'quantidade_atual', 'preco_medio', 'percentual_alvo', 'updated_at'])
						sucesso_formulario = 'Ativo atualizado com sucesso.'

		elif acao == 'excluir':
			ativo = get_object_or_404(Ativo, pk=request.POST.get('ativo_id'))
			ativo.delete()
			sucesso_formulario = 'Ativo excluído com sucesso.'

	tipo_filtro = (request.GET.get('tipo') or '').strip()
	nome_filtro = (request.GET.get('nome') or '').strip()
	setor_filtro = (request.GET.get('setor') or '').strip()
	mostrar_zerados = (request.GET.get('mostrar_zerados') == '1')

	ativos = Ativo.objects.all()
	if not mostrar_zerados:
		ativos = ativos.filter(quantidade_atual__gt=0)
	if tipo_filtro in tipos_validos:
		ativos = ativos.filter(tipo=tipo_filtro)
	if nome_filtro:
		ativos = ativos.filter(Q(nome__icontains=nome_filtro) | Q(ticker__icontains=nome_filtro))
	if setor_filtro:
		ativos = ativos.filter(setor__icontains=setor_filtro)

	ativos = ativos.order_by('ticker', 'nome')
	setores_disponiveis = (
		Ativo.objects.exclude(setor='')
		.values_list('setor', flat=True)
		.distinct()
		.order_by('setor')
	)

	context = {
		'ativos': ativos,
		'erro_formulario': erro_formulario,
		'sucesso_formulario': sucesso_formulario,
		'tipos_ativo': Ativo.Tipo.choices,
		'setores_disponiveis': setores_disponiveis,
		'filtros': {
			'tipo': tipo_filtro,
			'nome': nome_filtro,
			'setor': setor_filtro,
			'mostrar_zerados': mostrar_zerados,
		},
	}
	return render(request, 'investimentos/ativos.html', context)


@require_http_methods(["GET", "POST"])
def historico_investimentos(request):
	erro_formulario = None
	sucesso_formulario = None

	if request.method == 'POST':
		acao = request.POST.get('acao')
		if acao == 'editar_ordem':
			ordem = get_object_or_404(Ordem, pk=request.POST.get('item_id'))
			try:
				data_ordem = date.fromisoformat(request.POST.get('data') or '')
			except ValueError:
				data_ordem = None

			tipo = request.POST.get('tipo')
			quantidade = _to_decimal_raw(request.POST.get('quantidade')).quantize(Decimal('0.00000001'))
			preco = _to_decimal_raw(request.POST.get('preco')).quantize(Decimal('0.00000001'))
			taxas = _to_decimal_raw(request.POST.get('taxas')).quantize(Decimal('0.01'))

			if tipo not in [Ordem.TipoOrdem.COMPRA, Ordem.TipoOrdem.VENDA]:
				erro_formulario = 'Tipo de ordem inválido.'
			elif quantidade <= 0 or preco <= 0:
				erro_formulario = 'Quantidade e preço devem ser maiores que zero.'
			elif not data_ordem:
				erro_formulario = 'Data inválida para a ordem.'
			elif taxas < 0:
				erro_formulario = 'Taxas não podem ser negativas.'
			else:
				try:
					with transaction.atomic():
						ordem.tipo = tipo
						ordem.quantidade = quantidade
						ordem.preco = preco
						ordem.taxas = taxas
						ordem.data = data_ordem
						ordem.save(update_fields=['tipo', 'quantidade', 'preco', 'taxas', 'data', 'updated_at'])
						recalcular_posicao_ativo(ordem.ativo_id)
				except ValueError as exc:
					erro_formulario = str(exc)
				else:
					sucesso_formulario = 'Ordem atualizada com sucesso.'

		elif acao == 'editar_rendimento':
			rendimento = get_object_or_404(Rendimento, pk=request.POST.get('item_id'))
			try:
				data_rendimento = date.fromisoformat(request.POST.get('data') or '')
			except ValueError:
				data_rendimento = None

			valor = _to_decimal_raw(request.POST.get('valor')).quantize(Decimal('0.01'))
			descricao = (request.POST.get('descricao') or '').strip()

			if valor <= 0:
				erro_formulario = 'Valor do rendimento deve ser maior que zero.'
			elif not data_rendimento:
				erro_formulario = 'Data inválida para o rendimento.'
			else:
				rendimento.valor = valor
				rendimento.data = data_rendimento
				rendimento.descricao = descricao
				rendimento.save(update_fields=['valor', 'data', 'descricao', 'updated_at'])
				sucesso_formulario = 'Rendimento atualizado com sucesso.'

		elif acao == 'excluir_ordem':
			ordem = get_object_or_404(Ordem, pk=request.POST.get('item_id'))
			ativo_id = ordem.ativo_id
			with transaction.atomic():
				ordem.delete()
				recalcular_posicao_ativo(ativo_id)
			sucesso_formulario = 'Ordem excluída com sucesso.'

		elif acao == 'excluir_rendimento':
			rendimento = get_object_or_404(Rendimento, pk=request.POST.get('item_id'))
			rendimento.delete()
			sucesso_formulario = 'Rendimento excluído com sucesso.'

		elif acao == 'criar_aporte':
			valor = _to_decimal_raw(request.POST.get('valor')).quantize(Decimal('0.01'))
			descricao = (request.POST.get('descricao') or '').strip() or 'Aporte manual'
			data_raw = (request.POST.get('data') or '').strip()

			if valor <= 0:
				erro_formulario = 'Valor do aporte deve ser maior que zero.'
			else:
				aporte = AportePatrimonial.objects.create(
					valor=valor,
					descricao=descricao,
					id_transacao_origem=None,
				)
				if data_raw:
					try:
						data_aporte = date.fromisoformat(data_raw)
					except ValueError:
						erro_formulario = 'Data inválida para o aporte manual.'
					else:
						aporte.data = timezone.make_aware(datetime.combine(data_aporte, datetime.min.time()))
						aporte.save(update_fields=['data', 'updated_at'])
				if not erro_formulario:
					sucesso_formulario = 'Aporte manual cadastrado com sucesso.'

		elif acao == 'editar_aporte':
			aporte = get_object_or_404(AportePatrimonial, pk=request.POST.get('item_id'))
			valor = _to_decimal_raw(request.POST.get('valor')).quantize(Decimal('0.01'))
			descricao = (request.POST.get('descricao') or '').strip() or 'Aporte manual'
			data_raw = (request.POST.get('data') or '').strip()

			if valor <= 0:
				erro_formulario = 'Valor do aporte deve ser maior que zero.'
			else:
				aporte.valor = valor
				aporte.descricao = descricao
				if data_raw:
					try:
						data_aporte = date.fromisoformat(data_raw)
					except ValueError:
						erro_formulario = 'Data inválida para o aporte.'
					else:
						aporte.data = timezone.make_aware(datetime.combine(data_aporte, datetime.min.time()))
				if not erro_formulario:
					aporte.save(update_fields=['valor', 'descricao', 'data', 'updated_at'])
					sucesso_formulario = 'Aporte atualizado com sucesso.'

		elif acao == 'excluir_aporte':
			aporte = get_object_or_404(AportePatrimonial, pk=request.POST.get('item_id'))
			aporte.delete()
			sucesso_formulario = 'Aporte excluído com sucesso.'

	evento = (request.GET.get('evento') or '').strip().lower()
	if evento not in ['', 'ordem', 'rendimento', 'aporte']:
		evento = ''

	ativo_id = (request.GET.get('ativo_id') or '').strip()
	tipo_ordem = (request.GET.get('tipo_ordem') or '').strip()
	busca = (request.GET.get('busca') or '').strip()
	data_range = (request.GET.get('data_range') or '').strip()
	per_page_raw = (request.GET.get('per_page') or '30').strip()
	per_page = per_page_raw if per_page_raw in ['30', '50', '100'] else '30'
	inicio, fim = _parse_date_range(data_range)

	ativos = Ativo.objects.order_by('ticker', 'nome')

	ordens = Ordem.objects.select_related('ativo').order_by('-data', '-created_at')
	rendimentos = Rendimento.objects.select_related('ativo').order_by('-data', '-created_at')
	aportes = AportePatrimonial.objects.order_by('-data', '-created_at')

	if ativo_id:
		ordens = ordens.filter(ativo_id=ativo_id)
		rendimentos = rendimentos.filter(ativo_id=ativo_id)

	if tipo_ordem in [Ordem.TipoOrdem.COMPRA, Ordem.TipoOrdem.VENDA]:
		ordens = ordens.filter(tipo=tipo_ordem)

	if inicio and fim:
		ordens = ordens.filter(data__range=(inicio, fim))
		rendimentos = rendimentos.filter(data__range=(inicio, fim))
		aportes = aportes.filter(data__date__range=(inicio, fim))

	if busca:
		ordens = ordens.filter(Q(ativo__nome__icontains=busca) | Q(ativo__ticker__icontains=busca))
		rendimentos = rendimentos.filter(descricao__icontains=busca)
		aportes = aportes.filter(descricao__icontains=busca)

	eventos = []
	if evento in ['', 'ordem']:
		for ordem in ordens:
			eventos.append(
				{
					'id': ordem.id,
					'categoria': 'Ordem',
					'tipo': ordem.tipo,
					'data': ordem.data,
					'ativo': ordem.ativo,
					'ativo_label': (ordem.ativo.ticker or ordem.ativo.nome),
					'descricao': f'{ordem.get_tipo_display()} de {ordem.quantidade} @ R$ {ordem.preco}',
					'valor': ordem.total,
					'quantidade': ordem.quantidade,
					'preco': ordem.preco,
					'taxas': ordem.taxas,
					'created_at': ordem.created_at,
				}
			)

	if evento in ['', 'rendimento']:
		for rendimento in rendimentos:
			eventos.append(
				{
					'id': rendimento.id,
					'categoria': 'Rendimento',
					'tipo': 'Rendimento',
					'data': rendimento.data,
					'ativo': rendimento.ativo,
					'ativo_label': (rendimento.ativo.ticker or rendimento.ativo.nome),
					'descricao': rendimento.descricao or 'Provento registrado',
					'valor': rendimento.valor,
					'resgatar_para_orcamento': rendimento.resgatar_para_orcamento,
					'created_at': rendimento.created_at,
				}
			)

	if evento in ['', 'aporte']:
		for aporte in aportes:
			eventos.append(
				{
					'id': aporte.id,
					'categoria': 'Aporte',
					'tipo': 'Aporte',
					'data': aporte.data.date(),
					'ativo': None,
					'ativo_label': 'Patrimônio',
					'descricao': aporte.descricao,
					'valor': aporte.valor,
					'id_transacao_origem': aporte.id_transacao_origem,
					'created_at': aporte.created_at,
				}
			)

	eventos.sort(key=lambda item: (item['data'], item['created_at']), reverse=True)

	paginator = Paginator(eventos, int(per_page))
	page_number = request.GET.get('page')
	eventos_page = paginator.get_page(page_number)

	params = request.GET.copy()
	params.pop('page', None)
	page_query_prefix = params.urlencode()
	if page_query_prefix:
		page_query_prefix += '&'

	context = {
		'eventos': eventos_page,
		'ativos': ativos,
		'hoje': date.today(),
		'erro_formulario': erro_formulario,
		'sucesso_formulario': sucesso_formulario,
		'filtro_evento': evento,
		'filtros': {
			'evento': evento,
			'ativo_id': ativo_id,
			'tipo_ordem': tipo_ordem,
			'busca': busca,
			'data_range': data_range,
			'per_page': per_page,
		},
		'page_query_prefix': page_query_prefix,
	}
	return render(request, 'investimentos/historico.html', context)


@require_http_methods(["GET"])
def baixar_modelo_importacao_xlsx(request):
	try:
		from openpyxl import Workbook
	except ImportError:
		return HttpResponse('openpyxl não está instalado no ambiente.', status=500)

	wb = Workbook()
	ws = wb.active
	ws.title = 'dados'

	headers = [
		'tipo_registro',
		'ticker',
		'nome_ativo',
		'tipo_ativo',
		'tipo_ordem',
		'quantidade',
		'preco',
		'taxas',
		'valor_rendimento',
		'data',
		'descricao',
		'resgatar_para_orcamento',
	]
	ws.append(headers)

	ws.append(['Compra', 'WEGE3', 'WEG S.A.', 'Acao', 'Compra', 10, 40.00, 0, '', '2026-01-15', 'Compra inicial', 0])
	ws.append(['Venda', 'ITUB4', 'Itaú Unibanco', 'Acao', 'Venda', 2, 30.50, 0.50, '', '2026-02-10', 'Realização parcial', 0])
	ws.append(['Rendimento', 'WEGE3', 'WEG S.A.', 'Acao', '', '', '', '', 25.75, '2026-02-28', 'Dividendos', 1])

	response = HttpResponse(
		content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
	)
	response['Content-Disposition'] = 'attachment; filename="modelo_importacao_investimentos.xlsx"'
	wb.save(response)
	return response


@require_http_methods(["GET", "POST"])
def importar_operacoes_xlsx(request):
	erro_formulario = None
	sucesso_formulario = None
	linhas_com_erro = []
	registros_processados = 0

	if request.method == 'POST':
		arquivo = request.FILES.get('arquivo_xlsx')
		if not arquivo:
			erro_formulario = 'Selecione um arquivo .xlsx para importar.'
		elif not (arquivo.name or '').lower().endswith('.xlsx'):
			erro_formulario = 'Formato inválido. Envie um arquivo .xlsx.'
		else:
			try:
				from openpyxl import load_workbook
			except ImportError:
				erro_formulario = 'openpyxl não está instalado no ambiente.'
			else:
				workbook = load_workbook(filename=arquivo, data_only=True)
				sheet = workbook.active
				rows = list(sheet.iter_rows(values_only=True))

				if len(rows) <= 1:
					erro_formulario = 'O arquivo não possui dados para importação.'
				else:
					headers = [str(col or '').strip().lower() for col in rows[0]]
					mapa = {header: idx for idx, header in enumerate(headers) if header}

					for numero_linha, valores in enumerate(rows[1:], start=2):
						if not any(valor not in [None, ''] for valor in valores):
							continue

						try:
							tipo_registro = str(valores[mapa.get('tipo_registro', -1)] or '').strip().lower()
							ticker = str(valores[mapa.get('ticker', -1)] or '').strip().upper()
							nome_ativo = str(valores[mapa.get('nome_ativo', -1)] or '').strip()
							tipo_ativo = str(valores[mapa.get('tipo_ativo', -1)] or '').strip() or Ativo.Tipo.ACAO
							data_raw = valores[mapa.get('data', -1)]
							descricao = str(valores[mapa.get('descricao', -1)] or '').strip()

							if hasattr(data_raw, 'date'):
								data_ref = data_raw.date() if hasattr(data_raw, 'hour') else data_raw
							else:
								data_ref = date.fromisoformat(str(data_raw or '').strip())

							ativo = _resolve_or_create_ativo(ticker, nome_ativo, tipo_ativo)

							if tipo_registro in {'compra', 'venda', 'ordem'}:
								tipo_ordem = str(valores[mapa.get('tipo_ordem', -1)] or '').strip() or tipo_registro.title()
								if tipo_ordem not in [Ordem.TipoOrdem.COMPRA, Ordem.TipoOrdem.VENDA]:
									raise ValueError('tipo_ordem inválido para registro de ordem.')

								quantidade = _to_decimal_raw(valores[mapa.get('quantidade', -1)]).quantize(Decimal('0.00000001'))
								preco = _to_decimal_raw(valores[mapa.get('preco', -1)]).quantize(Decimal('0.00000001'))
								taxas = _to_decimal_raw(valores[mapa.get('taxas', -1)]).quantize(Decimal('0.01'))

								if quantidade <= 0 or preco <= 0:
									raise ValueError('quantidade/preço devem ser maiores que zero.')

								ordem = Ordem.objects.create(
									ativo=ativo,
									tipo=tipo_ordem,
									quantidade=quantidade,
									preco=preco,
									taxas=taxas,
									data=data_ref,
								)
								processar_ordem(ordem)

							elif tipo_registro == 'rendimento':
								valor = _to_decimal_raw(valores[mapa.get('valor_rendimento', -1)]).quantize(Decimal('0.01'))
								if valor <= 0:
									raise ValueError('valor_rendimento deve ser maior que zero.')

								rendimento = Rendimento.objects.create(
									ativo=ativo,
									valor=valor,
									data=data_ref,
									descricao=descricao,
									resgatar_para_orcamento=_parse_bool_cell(valores[mapa.get('resgatar_para_orcamento', -1)]),
								)
								processar_rendimento(rendimento)
							else:
								raise ValueError('tipo_registro deve ser Compra, Venda ou Rendimento.')

							registros_processados += 1

						except Exception as exc:
							linhas_com_erro.append({'linha': numero_linha, 'erro': str(exc)})

					if registros_processados > 0:
						sucesso_formulario = f'{registros_processados} registro(s) importado(s) com sucesso.'
					if not sucesso_formulario and linhas_com_erro:
						erro_formulario = 'Nenhum registro foi importado. Revise os erros da planilha.'

	context = {
		'erro_formulario': erro_formulario,
		'sucesso_formulario': sucesso_formulario,
		'linhas_com_erro': linhas_com_erro,
	}
	return render(request, 'investimentos/importador_xlsx.html', context)
